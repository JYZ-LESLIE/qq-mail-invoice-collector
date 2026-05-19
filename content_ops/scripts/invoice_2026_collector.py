#!/usr/bin/env python3
"""Read-only QQ mailbox invoice collector for 2026 invoices.

The collector is conservative:
- read mail through IMAP only
- download likely invoice attachments
- recover invoice PDFs/OFDs from message links when possible
- record link/QR evidence for manual review when automated recovery is unsafe
- never send, delete, or move mail
"""

from __future__ import annotations

import argparse
import base64
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv
import datetime as dt
import email
from email.header import decode_header
from email.message import Message
import hashlib
import html
from html.parser import HTMLParser
import io
import imaplib
import json
import math
import os
from pathlib import Path
import re
import shutil
import sqlite3
import ssl
import threading
import time
from typing import Callable, Iterable, TypeVar
import xml.etree.ElementTree as ET
import zipfile
from urllib import request as urlrequest
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qsl, unquote, urljoin, urlparse

from invoice_pdf_parser import (
    build_mimo_qr_url_extractor,
    build_mimo_vision_extractor,
    has_minimum_invoice_fields,
    parse_invoice_pdf_pages,
    parse_invoice_fields_from_text,
    should_reject_as_non_invoice_document,
    split_pdf_page_to_bytes,
)


WORKSPACE_ROOT = Path(__file__).resolve().parents[2]
INVOICE_ROOT = WORKSPACE_ROOT / "发票整理"
CONFIG_DIR = INVOICE_ROOT / "私密配置"
RAW_DIR = INVOICE_ROOT / "原始附件"
PROCESSED_DIR = INVOICE_ROOT / "已整理发票"
MANUAL_DIR = INVOICE_ROOT / "人工复核"
REPORT_DIR = INVOICE_ROOT / "台账"
STATE_DIR = INVOICE_ROOT / "运行状态"
DEFAULT_LOG_DB = STATE_DIR / "log.db"
MAX_WORKERS = 3
OUTPUT_LOCK = threading.Lock()
THREAD_LOCAL = threading.local()

SUBJECT_KEYWORDS = (
    "发票",
    "电子发票",
    "电票",
    "数电票",
    "Invoice",
    "账单",
    "行程单",
    "滴滴",
    "小桔",
    "曹操出行",
    "高德打车",
    "T3出行",
    "首汽约车",
    "出行发票",
    "行程报销凭证",
)

PRIORITY_SENDER_DOMAINS = (
    "didi.com",
    "didichuxing.com",
    "didiglobal.com",
    "xiaojukeji.com",
    "udache.com",
    "diditaxi.com.cn",
    "caocaokeji.cn",
    "caocaomobility.com",
    "t3go.cn",
    "01zhuanche.com",
    "amap.com",
    "autonavi.com",
    "meituan.com",
    "dianping.com",
    "trip.com",
    "ctrip.com",
    "qunar.com",
    "12306.cn",
    "railway12306.cn",
    "apple.com",
    "stripe.com",
)

LIKELY_INVOICE_KEYWORDS = (
    "发票",
    "电子发票",
    "数电票",
    "全电票",
    "开票",
    "电票",
    "数电票",
    "票据",
    "行程单",
    "报销",
    "行程报销凭证",
    "打车发票",
    "用车发票",
    "出行发票",
    "电子行程单",
    "网约车",
    "曹操出行",
    "高德打车",
    "t3出行",
    "首汽约车",
    "invoice",
    "receipt",
    "tax",
)

DOWNLOAD_ATTACHMENT_EXTENSIONS = {".pdf", ".ofd"}
SKIPPED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tif", ".tiff"}
MIN_IMAGE_SIGNATURE_BYTES = 20 * 1024
INVOICE_EXTENSIONS = {".pdf", ".ofd", ".xml"}
ARCHIVE_EXTENSIONS = {".zip"}
URL_RE = re.compile(r"https?://[^\s\"'<>）)]+", re.IGNORECASE)
LINK_DOWNLOAD_TIMEOUT_SECONDS = 30
MAX_LINK_DOWNLOAD_BYTES = 25 * 1024 * 1024
MAX_LINK_CANDIDATES_PER_MAIL = 20
MAX_ATTACHMENT_PROBE_UIDS = 500
BANK_STATEMENT_BRANDS = (
    "招商银行",
    "招商银行信用卡",
    "民生银行",
    "民生信用卡",
    "中信银行",
    "中信银行信用卡",
    "中国工商银行",
    "工商银行",
    "icbc",
    "za bank",
    "bochk",
)
BANK_STATEMENT_HINTS = (
    "对账单",
    "电子对账单",
    "电子账单",
    "银行账单",
    "信用卡账单",
    "信用卡电子账单",
    "客户对账单",
    "bank statement",
    "statement",
    "结单",
)
BANK_SENDER_HINTS = (
    "creditcard.cmbc.com.cn",
    "message.cmbchina.com",
    "icbc.com.cn",
    "citicbank.com",
    "bank.ecitic.com",
    "za.group",
    "za.bank",
    "bochk.com",
)
NON_REIMBURSABLE_MARKERS = (
    "netlify, inc",
    "invoices.withorb.com",
    "出行确认单（附件：中英文行程单）",
    "中英文行程单",
    "ctrip_booking_confirmed",
)
NON_INVOICE_DOCUMENT_MARKERS = (
    "报价编号",
    "报价单",
    "报价有效期",
    "预计总价",
    "quotation",
    "quote number",
    "proforma invoice",
)
DEFERRED_PLATFORM_MARKERS = (
    "icloud+",
    "icloud-efapiao.gzdata.com.cn",
)
CTRIP_UTILITY_PATH_MARKERS = (
    "icp.pdf",
    "internetdrugcertificate.pdf",
    "yiliao.pdf",
    "wxsb/",
)
LINK_INVOICE_HINTS = (
    "发票",
    "下载",
    "查看",
    "获取",
    "开票",
    "打开链接",
    "pdf",
    "ofd",
    "xml",
    "invoice",
    "receipt",
    "fapiao",
    "download",
    "bill",
    "tax",
)
LINK_NOISE_HINTS = (
    "unsubscribe",
    "privacy",
    "terms",
    "legal",
    "support",
    "help",
    "track",
    "tracking",
    "preference",
    "logo",
    "banner",
)
BWJF_TRACKING_HOSTS = {
    "bdopcs.bwjf.cn",
}
BWJF_TRACKING_PATH_MARKERS = (
    "/v1/usereventtransforget",
)
BWJF_DOWNLOAD_URL_KEYS = (
    "pdfurl",
    "pdf_url",
)
DECORATIVE_IMAGE_NAME_HINTS = (
    "logo",
    "banner",
    "icon",
    "footer",
    "header",
    "signature",
    "avatar",
    "sprite",
)
QR_HINTS = ("二维码", "扫码", "扫一扫", "QR", "qr", "qrcode", "领取发票", "下载发票")
T = TypeVar("T")


class _AnchorLinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.links: list[tuple[str, str]] = []
        self._current_href = ""
        self._current_text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        attr_map = {key.lower(): value or "" for key, value in attrs}
        self._current_href = attr_map.get("href", "")
        self._current_text = []

    def handle_data(self, data: str) -> None:
        if self._current_href:
            self._current_text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() != "a" or not self._current_href:
            return
        anchor_text = re.sub(r"\s+", " ", "".join(self._current_text)).strip()
        self.links.append((self._current_href, anchor_text))
        self._current_href = ""
        self._current_text = []


class _ImageSourceParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.images: list[dict[str, str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "img":
            return
        attr_map = {key.lower(): value or "" for key, value in attrs}
        src = attr_map.get("src", "").strip()
        if not src:
            return
        self.images.append(
            {
                "src": src,
                "alt": attr_map.get("alt", ""),
                "title": attr_map.get("title", ""),
            }
        )


class PlainProgress:
    def __init__(self, total: int, desc: str = ""):
        self.total = total
        self.desc = desc
        self.count = 0

    def __enter__(self) -> "PlainProgress":
        if self.desc:
            print(f"{self.desc}: 0/{self.total}")
        return self

    def __exit__(self, exc_type: object, exc: object, tb: object) -> None:
        pass

    def update(self, value: int = 1) -> None:
        self.count += value
        if self.count == self.total or self.count % 25 == 0:
            print(f"{self.desc}: {self.count}/{self.total}")

    def set_postfix(self, *_args: object, **_kwargs: object) -> None:
        pass


def make_progress(total: int, desc: str):
    try:
        from tqdm import tqdm  # type: ignore

        return tqdm(total=total, desc=desc, unit="mail")
    except Exception:
        return PlainProgress(total=total, desc=desc)


def retry_call(
    func: Callable[[], T],
    *,
    attempts: int = 3,
    sleep_seconds: float = 2.0,
    label: str = "operation",
) -> T:
    last_exc: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            return func()
        except Exception as exc:
            last_exc = exc
            if attempt >= attempts:
                break
            print(f"{label} failed on attempt {attempt}/{attempts}; retrying...")
            time.sleep(sleep_seconds * attempt)
    assert last_exc is not None
    raise last_exc


def load_env(path: Path | None) -> dict[str, str]:
    values: dict[str, str] = {}
    if path and path.exists():
        for raw_line in path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            values[key.strip()] = value.strip().strip('"').strip("'")
    merged = dict(os.environ)
    merged.update(values)
    return merged


def decode_mime_words(value: str | None) -> str:
    if not value:
        return ""
    if not isinstance(value, str):
        value = str(value)
    parts = []
    for data, charset in decode_header(value):
        if isinstance(data, bytes):
            charset = charset or "utf-8"
            try:
                parts.append(data.decode(charset, errors="replace"))
            except LookupError:
                parts.append(data.decode("utf-8", errors="replace"))
        else:
            parts.append(data)
    return "".join(parts).strip()


def safe_name(value: str, default: str = "unnamed") -> str:
    text = decode_mime_words(value) or default
    text = html.unescape(text)
    text = re.sub(r"[\\/:*?\"<>|\r\n\t]+", "_", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text[:160] or default


def parse_date(value: str | None) -> str:
    if not value:
        return ""
    value_text = decode_mime_words(value)
    try:
        parsed = email.utils.parsedate_to_datetime(value_text)
        if parsed.tzinfo:
            parsed = parsed.astimezone(dt.timezone(dt.timedelta(hours=8)))
        return parsed.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return value_text[:80]


def date_to_imap(value: str) -> str:
    parsed = dt.datetime.strptime(value, "%Y-%m-%d").date()
    return parsed.strftime("%d-%b-%Y")


def next_date(value: str) -> str:
    parsed = dt.datetime.strptime(value, "%Y-%m-%d").date()
    return (parsed + dt.timedelta(days=1)).isoformat()


def part_filename(part: Message) -> str:
    filename = part.get_filename()
    if filename:
        return safe_name(filename)
    content_type = part.get_content_type()
    ext = {
        "application/pdf": ".pdf",
        "application/xml": ".xml",
        "text/xml": ".xml",
        "application/zip": ".zip",
        "image/jpeg": ".jpg",
        "image/png": ".png",
    }.get(content_type, "")
    return f"attachment{ext}"


def encoded_payload_size(part: Message) -> int:
    size_value = part.get_param("size", header="content-disposition")
    if size_value:
        try:
            return int(str(size_value))
        except ValueError:
            pass
    payload = part.get_payload(decode=False)
    if isinstance(payload, bytes):
        return len(payload)
    if isinstance(payload, str):
        compact = re.sub(r"\s+", "", payload)
        return int(math.ceil(len(compact) * 0.75))
    return 0


def should_download_attachment(part: Message, filename: str) -> bool:
    name = filename or part_filename(part)
    ext = Path(name).suffix.lower()
    content_type = (part.get_content_type() or "").lower()
    estimated_size = encoded_payload_size(part)

    if content_type.startswith("image/") or ext in SKIPPED_IMAGE_EXTENSIONS:
        if estimated_size and estimated_size < MIN_IMAGE_SIGNATURE_BYTES:
            return False
        if not any(token in name for token in ("发票", "Invoice", "invoice")):
            return False

    return ext in DOWNLOAD_ATTACHMENT_EXTENSIONS or "发票" in name or "Invoice" in name or "invoice" in name


def message_has_invoice_attachment(msg: Message) -> bool:
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        disposition = str(part.get("Content-Disposition") or "").lower()
        filename = part.get_filename()
        if "attachment" not in disposition and not filename:
            continue
        if should_download_attachment(part, part_filename(part)):
            return True
    return False


def get_text_parts(msg: Message) -> tuple[str, str]:
    plain_parts: list[str] = []
    html_parts: list[str] = []
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        content_type = part.get_content_type()
        if content_type not in {"text/plain", "text/html"}:
            continue
        payload = part.get_payload(decode=True)
        if payload is None:
            continue
        charset = part.get_content_charset() or "utf-8"
        try:
            text = payload.decode(charset, errors="replace")
        except LookupError:
            text = payload.decode("utf-8", errors="replace")
        if content_type == "text/html":
            html_parts.append(text)
        else:
            plain_parts.append(text)
    return "\n".join(plain_parts), "\n".join(html_parts)


def looks_invoice_related(*texts: str) -> bool:
    combined = " ".join(t or "" for t in texts).lower()
    return any(keyword.lower() in combined for keyword in LIKELY_INVOICE_KEYWORDS)


def subject_has_invoice_keyword(subject: str) -> bool:
    subject_lower = (subject or "").lower()
    return any(keyword.lower() in subject_lower for keyword in SUBJECT_KEYWORDS)


def invoice_hints_from_subject(subject: str) -> dict[str, str]:
    hints: dict[str, str] = {}
    text = subject or ""
    seller_patterns = (
        r"您收到一张【(.+?)】开具的",
        r"您收到来自(.+?)的电子发票",
        r"【电子发票】\s*(.+?)（发票金额",
        r"电子发票下载\s+(.+?)-[0-9]{8,}",
    )
    for pattern in seller_patterns:
        match = re.search(pattern, text)
        if match:
            hints["seller"] = match.group(1).strip()
            break
    number_match = re.search(r"发票号码[：:]?\s*([0-9A-Za-z-]+)", text)
    if number_match:
        hints["invoice_number"] = number_match.group(1).strip()
    amount_match = re.search(r"发票金额[：:]?\s*([0-9,]+(?:\.[0-9]{1,2})?)", text)
    if amount_match:
        hints["amount"] = f"{float(amount_match.group(1).replace(',', '')):.2f}"
    return hints


def is_bank_statement_email(subject: str, sender: str, body_text: str = "") -> bool:
    header_text = f"{subject} {sender}".lower()
    combined = f"{header_text} {body_text[:1500]}".lower()
    has_bank_sender = any(hint in header_text for hint in BANK_SENDER_HINTS)
    has_bank_brand = any(brand.lower() in combined for brand in BANK_STATEMENT_BRANDS)
    has_statement_hint = any(hint.lower() in combined for hint in BANK_STATEMENT_HINTS)
    return has_statement_hint and (has_bank_sender or has_bank_brand)


def is_bank_statement_row(row: dict[str, str]) -> bool:
    return is_bank_statement_email(
        str(row.get("mail_subject") or ""),
        str(row.get("mail_from") or ""),
        " ".join(
            str(row.get(key) or "")
            for key in (
                "seller",
                "attachment_original_name",
                "original_file",
                "output_file",
                "links",
                "note",
            )
        ),
    )


def is_non_reimbursable_email(subject: str, sender: str, body_text: str = "") -> bool:
    combined = f"{subject} {sender} {body_text[:2500]}".lower()
    if any(marker.lower() in combined for marker in NON_REIMBURSABLE_MARKERS):
        return True
    if any(marker in combined for marker in CTRIP_UTILITY_PATH_MARKERS):
        return True
    return False


def is_deferred_platform_email(subject: str, sender: str, body_text: str = "") -> bool:
    combined = f"{subject} {sender} {body_text[:2500]}".lower()
    return any(marker.lower() in combined for marker in DEFERRED_PLATFORM_MARKERS)


def is_excluded_business_row(row: dict[str, str]) -> bool:
    combined = " ".join(
        str(row.get(key) or "")
        for key in (
            "mail_subject",
            "mail_from",
            "seller",
            "purchaser",
            "attachment_original_name",
            "original_file",
            "output_file",
            "links",
            "note",
            "link_platform",
        )
    ).lower()
    if any(marker.lower() in combined for marker in NON_REIMBURSABLE_MARKERS):
        return True
    if any(marker.lower() in combined for marker in DEFERRED_PLATFORM_MARKERS):
        return True
    if any(marker.lower() in combined for marker in NON_INVOICE_DOCUMENT_MARKERS):
        return True
    output_text = extract_pdf_text_for_row(row)
    if output_text and should_reject_as_non_invoice_document(output_text):
        return True
    if any(marker in combined for marker in CTRIP_UTILITY_PATH_MARKERS):
        return True
    return False


def extract_pdf_text_for_row(row: dict[str, str]) -> str:
    path_text = str(row.get("output_file") or row.get("original_file") or "")
    if not path_text:
        return ""
    path = Path(path_text)
    if path.suffix.lower() != ".pdf" or not path.exists():
        return ""
    try:
        import fitz  # type: ignore

        with fitz.open(path) as doc:
            return "\n".join(doc.load_page(i).get_text("text") or "" for i in range(min(len(doc), 2)))
    except Exception:
        return ""


def output_is_pdf(row: dict[str, str]) -> bool:
    output_file = str(row.get("output_file") or row.get("original_file") or row.get("attachment_original_name") or "")
    return Path(output_file).suffix.lower() == ".pdf"


def looks_like_china_invoice_row(row: dict[str, str]) -> bool:
    combined = " ".join(
        str(row.get(key) or "")
        for key in (
            "invoice_kind",
            "category",
            "mail_subject",
            "attachment_original_name",
            "original_file",
            "output_file",
            "seller",
            "purchaser",
            "note",
        )
    )
    markers = ("增值税", "电子发票", "数电", "普通发票", "专用发票", "发票号码", "开票日期")
    has_tax_marker = any(marker in combined for marker in markers)
    invoice_number = re.sub(r"\D", "", str(row.get("invoice_number") or ""))
    has_cn_name = bool(re.search(r"[\u4e00-\u9fff]{2,}", f"{row.get('seller', '')}{row.get('purchaser', '')}"))
    if invoice_number and len(invoice_number) >= 8 and (has_cn_name or has_tax_marker):
        return True
    if has_tax_marker and has_cn_name and row.get("amount") and row.get("invoice_date"):
        return True
    return False


def amount_to_float(value: object) -> float | None:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def effective_amount(row: dict[str, str]) -> float:
    value = amount_to_float(row.get("effective_amount"))
    if value is not None:
        return value
    value = amount_to_float(row.get("amount"))
    return value or 0.0


def row_is_countable_invoice(row: dict[str, str]) -> bool:
    if row.get("status") not in {"Parsed", "AI_Verified"}:
        return False
    return str(row.get("include_in_summary") or "是").strip() != "否"


def extract_order_number_from_text(*values: object) -> str:
    text = " ".join(str(value or "") for value in values)
    match = re.search(r"(W\d{10})(?!\d)", text, flags=re.IGNORECASE)
    return match.group(1).upper() if match else ""


def extract_red_blue_invoice_number_from_text(*values: object) -> str:
    text = " ".join(str(value or "") for value in values)
    match = re.search(r"被红冲蓝字(?:数电)?发票号码[:：]?\s*([0-9]{8,30})", text)
    return match.group(1) if match else ""


def enrich_row_from_pdf(row: dict[str, str]) -> dict[str, str]:
    path_text = str(row.get("output_file") or row.get("original_file") or "")
    if not path_text:
        return row
    path = Path(path_text)
    if path.suffix.lower() != ".pdf" or not path.exists():
        return row
    needs_pdf_probe = (
        not row.get("order_number")
        or not row.get("related_invoice_number")
        or str(row.get("status") or "") not in {"Parsed", "AI_Verified"}
    )
    if not needs_pdf_probe:
        return row
    try:
        parsed_pages = parse_invoice_pdf_pages(path.read_bytes(), filename=path.name, vision_extract=None)
    except Exception:
        return row
    if not parsed_pages:
        return row
    parsed = parsed_pages[0]
    fields = {
        key: str(parsed.get(key, "") or "")
        for key in (
            "invoice_date",
            "seller",
            "purchaser",
            "amount",
            "invoice_code",
            "invoice_number",
            "invoice_kind",
            "category",
            "order_number",
            "related_invoice_number",
            "relation_type",
        )
    }
    for key, value in fields.items():
        if value and not row.get(key):
            row[key] = value
    if row.get("status") not in {"Parsed", "AI_Verified"} and has_minimum_invoice_fields(fields):
        row["status"] = "Parsed"
        row["parse_status"] = str(parsed.get("status") or "parsed")
        row["parse_engine"] = str(parsed.get("engine") or "pdf_text_repair")
        append_note_once(row, "auto_repaired_from_pdf_text")
    return row


def relation_defaults(row: dict[str, str]) -> None:
    amount = amount_to_float(row.get("amount"))
    if not row.get("order_number"):
        row["order_number"] = extract_order_number_from_text(row.get("mail_subject"), row.get("links"), row.get("attachment_original_name"))
    if not row.get("relation_type"):
        row["relation_type"] = ""
    if not row.get("relation_group"):
        row["relation_group"] = ""
    if not row.get("related_invoice_number"):
        row["related_invoice_number"] = ""
    if not row.get("include_in_summary"):
        row["include_in_summary"] = "是"
    if not row.get("effective_amount"):
        row["effective_amount"] = f"{amount:.2f}" if amount is not None else ""


def append_note_once(row: dict[str, str], text: str) -> None:
    current = str(row.get("note") or "")
    if text and text not in current:
        row["note"] = "; ".join(part for part in (current, text) if part)


def is_apple_hardware_invoice_row(row: dict[str, str]) -> bool:
    text = " ".join(
        str(row.get(key) or "")
        for key in ("seller", "mail_subject", "mail_from", "link_platform", "attachment_original_name", "output_file")
    ).lower()
    return "apple" in text or "苹果电子产品商贸" in text


def is_company_purchaser(name: object) -> bool:
    text = str(name or "")
    return any(token in text for token in ("公司", "有限", "企业", "中心", "事务所", "工作室"))


def annotate_invoice_relationships(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    annotated = [dict(row) for row in rows]
    for row in annotated:
        relation_defaults(row)
        if is_apple_hardware_invoice_row(row) or "联通" in str(row.get("seller") or row.get("mail_subject") or ""):
            enrich_row_from_pdf(row)
        if not row.get("order_number"):
            row["order_number"] = extract_order_number_from_text(row.get("mail_subject"), row.get("links"), row.get("attachment_original_name"), row.get("note"))
        if not row.get("related_invoice_number"):
            row["related_invoice_number"] = extract_red_blue_invoice_number_from_text(row.get("note"), row.get("links"), row.get("mail_subject"))
        amount = amount_to_float(row.get("amount"))
        if amount is not None:
            row["effective_amount"] = f"{amount:.2f}"
        if amount is not None and amount < 0:
            related = str(row.get("related_invoice_number") or "").strip()
            row["relation_type"] = "红字冲销"
            row["relation_group"] = f"red:{related or row.get('invoice_number') or row.get('source_uid')}"
            row["include_in_summary"] = "是"
            row["effective_amount"] = f"{amount:.2f}"
            if related:
                append_note_once(row, f"冲红蓝字发票 {related}")

    apple_groups: dict[str, list[dict[str, str]]] = {}
    for row in annotated:
        if row.get("status") not in {"Parsed", "AI_Verified"}:
            continue
        if not is_apple_hardware_invoice_row(row):
            continue
        order = str(row.get("order_number") or "").strip().upper()
        if order:
            apple_groups.setdefault(order, []).append(row)

    for order, group in apple_groups.items():
        company_rows = [row for row in group if is_company_purchaser(row.get("purchaser"))]
        personal_rows = [row for row in group if row not in company_rows]
        if not company_rows or not personal_rows:
            continue
        company_numbers = " / ".join(dict.fromkeys(str(row.get("invoice_number") or "") for row in company_rows if row.get("invoice_number")))
        personal_numbers = " / ".join(dict.fromkeys(str(row.get("invoice_number") or "") for row in personal_rows if row.get("invoice_number")))
        group_id = f"reissue:apple:{order}"
        for row in personal_rows:
            row["relation_type"] = "换开原票"
            row["relation_group"] = group_id
            row["related_invoice_number"] = company_numbers
            row["include_in_summary"] = "否"
            row["effective_amount"] = "0.00"
            append_note_once(row, f"已换开为公司抬头发票 {company_numbers}")
        for row in company_rows:
            row["relation_type"] = "换开后有效发票"
            row["relation_group"] = group_id
            row["related_invoice_number"] = personal_numbers
            row["include_in_summary"] = "是"
            amount = amount_to_float(row.get("amount"))
            row["effective_amount"] = f"{(amount or 0.0):.2f}" if amount is not None else ""

    return annotated


def dedupe_invoice_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    best_by_key: dict[tuple[str, ...], tuple[int, int]] = {}

    def row_score(row: dict[str, str]) -> int:
        score = 0
        if output_is_pdf(row):
            score += 20
        if row.get("invoice_number"):
            score += 10
        if row.get("seller"):
            score += 5
        if row.get("status") == "Parsed":
            score += 3
        if row.get("status") == "AI_Verified":
            score += 2
        if row.get("parse_status") == "parsed":
            score += 1
        return score

    def row_dedupe_key(row: dict[str, str]) -> tuple[str, ...] | None:
        invoice_number = str(row.get("invoice_number") or "").strip()
        if invoice_number:
            return ("invoice_number", invoice_number)
        invoice_date = str(row.get("invoice_date") or "").strip()
        amount = str(row.get("amount") or "").strip()
        seller = str(row.get("seller") or "").strip()
        purchaser = str(row.get("purchaser") or "").strip()
        account_id = str(row.get("account_id") or "").strip()
        if invoice_date and amount and seller and purchaser:
            return ("date_amount_parties_without_number", invoice_date, amount, seller, purchaser, account_id)
        return None

    for index, row in enumerate(rows):
        if row.get("status") not in {"Parsed", "AI_Verified"}:
            continue
        key = row_dedupe_key(row)
        if not key:
            continue
        score = row_score(row)
        if key not in best_by_key or score > best_by_key[key][0]:
            best_by_key[key] = (score, index)

    keep_indexes = {best_index for _, best_index in best_by_key.values()}
    deduped: list[dict[str, str]] = []
    for index, row in enumerate(rows):
        if row.get("status") in {"Parsed", "AI_Verified"}:
            key = row_dedupe_key(row)
            if key and index not in keep_indexes:
                continue
        deduped.append(row)
    return deduped


def dedupe_review_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[tuple[str, ...]] = set()
    deduped: list[dict[str, str]] = []
    for row in rows:
        if row.get("status") in {"Parsed", "AI_Verified"}:
            deduped.append(row)
            continue
        key = (
            str(row.get("status") or ""),
            str(row.get("source_uid") or ""),
            str(row.get("mail_subject") or ""),
            str(row.get("page_number") or ""),
            str(row.get("invoice_number") or ""),
            str(row.get("amount") or ""),
            str(row.get("note") or ""),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def clean_manifest_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    rows = annotate_invoice_relationships(rows)
    filtered: list[dict[str, str]] = []
    for row in rows:
        status = str(row.get("status") or "")
        if is_bank_statement_row(row):
            continue
        if is_excluded_business_row(row):
            continue
        if status in {"Non_CN_Invoice", "Non_Invoice_Document"}:
            continue
        if status in {"Parsed", "AI_Verified"} and not looks_like_china_invoice_row(row):
            continue
        if status in {"Downloaded_Raw_Attachment", "Link_Downloaded_Raw"} and not output_is_pdf(row):
            continue
        filtered.append(row)
    filtered = dedupe_invoice_rows(filtered)
    filtered = annotate_invoice_relationships(filtered)
    parsed_numbers = {
        str(row.get("invoice_number") or "").strip()
        for row in filtered
        if row.get("status") in {"Parsed", "AI_Verified"} and str(row.get("invoice_number") or "").strip()
    }
    parsed_date_amount = {
        (str(row.get("invoice_date") or "").strip(), str(row.get("amount") or "").strip())
        for row in filtered
        if row.get("status") in {"Parsed", "AI_Verified"}
        and not str(row.get("invoice_number") or "").strip()
        and str(row.get("invoice_date") or "").strip()
        and str(row.get("amount") or "").strip()
    }
    parsed_uids = {
        str(row.get("source_uid") or "").strip()
        for row in filtered
        if row.get("status") in {"Parsed", "AI_Verified"} and str(row.get("source_uid") or "").strip()
    }
    final_rows: list[dict[str, str]] = []
    for row in filtered:
        if row.get("status") not in {"Parsed", "AI_Verified"}:
            invoice_number = str(row.get("invoice_number") or "").strip()
            invoice_date = str(row.get("invoice_date") or "").strip()
            amount = str(row.get("amount") or "").strip()
            uid = str(row.get("source_uid") or "").strip()
            subject = str(row.get("mail_subject") or "")
            links = str(row.get("links") or "").lower()
            if invoice_number and invoice_number in parsed_numbers:
                continue
            if not invoice_number and invoice_date and amount and (invoice_date, amount) in parsed_date_amount:
                continue
            artifact_number_match = re.search(
                r"\b(\d{20})\b",
                " ".join(
                    str(row.get(key) or "")
                    for key in ("attachment_original_name", "original_file", "output_file", "links")
                ),
            )
            if row.get("status") == "Need_Review" and artifact_number_match and artifact_number_match.group(1) in parsed_numbers:
                continue
            if uid and uid in parsed_uids and row.get("status") in {"Link_Need_Manual", "QR_Need_Manual"}:
                continue
            if uid and uid in parsed_uids and row.get("status") == "Need_Review" and not invoice_number:
                continue
            if row.get("status") == "Link_Need_Manual" and not row.get("output_file"):
                if "store.apple.com/us/go/app" in links:
                    continue
                if "Apple 订单" in subject and "fdfinvoice.com" in links:
                    continue
                if "淘宝闪购平台订单发票开具完成通知" in subject and ".zip" in links:
                    continue
        final_rows.append(row)
    return dedupe_review_rows(final_rows)


def extract_urls(*texts: str) -> list[str]:
    seen = set()
    urls = []
    for text in texts:
        for match in URL_RE.findall(html.unescape(text or "")):
            clean = match.rstrip(".,;，。；")
            if clean not in seen:
                seen.add(clean)
                urls.append(clean)
    return urls


def normalize_invoice_link(url: str, base_url: str = "") -> str:
    value = html.unescape(str(url or "")).strip()
    if not value or value.startswith(("mailto:", "tel:", "javascript:")):
        return ""
    if base_url:
        value = urljoin(base_url, value)
    parsed = urlparse(value)
    if parsed.scheme.lower() not in {"http", "https"}:
        return ""
    for key, query_value in parse_qsl(parsed.query, keep_blank_values=True):
        if key.lower() in {"pdfurl", "url", "target", "redirect"} and query_value.startswith(("http://", "https://")):
            return unquote(query_value).strip()
    return value.rstrip(".,;，。；")


def is_bwjf_tracking_link(url: str) -> bool:
    parsed = urlparse(str(url or ""))
    host = (parsed.hostname or "").lower()
    path = (parsed.path or "").lower()
    if host in BWJF_TRACKING_HOSTS:
        return True
    return any(marker in path for marker in BWJF_TRACKING_PATH_MARKERS)


def preferred_bwjf_pdf_url(url: str) -> str:
    parsed = urlparse(str(url or ""))
    for key, query_value in parse_qsl(parsed.query, keep_blank_values=True):
        if key.lower() in BWJF_DOWNLOAD_URL_KEYS and query_value.startswith(("http://", "https://")):
            return unquote(query_value).strip()
    return ""


def is_bwjf_pdf_download_url(url: str) -> bool:
    parsed = urlparse(str(url or ""))
    host = (parsed.hostname or "").lower()
    path = (parsed.path or "").lower()
    query = parsed.query.lower()
    if host != "fp.bwjf.cn":
        return False
    if path.startswith("/downsigninvoice"):
        return "jflx=ofd" not in query and "jflx=xml" not in query
    return path.lower().endswith(".pdf")


def is_bwjf_invoice_link(url: str) -> bool:
    if is_bwjf_tracking_link(url):
        return False
    parsed = urlparse(str(url or ""))
    host = (parsed.hostname or "").lower()
    path = (parsed.path or "").lower()
    if host == "fp.bwjf.cn" and (path.startswith("/u/") or path.startswith("/downsigninvoice")):
        return True
    if host == "www.bwjf.cn" and ("alleledeliverysuccess" in path or preferred_bwjf_pdf_url(url)):
        return True
    if host == "yjts.bwjf.cn":
        return True
    return False


def extract_link_candidates(plain: str, html_text: str) -> list[dict[str, str]]:
    candidates: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    seen_urls: set[str] = set()

    def add(url: str, anchor_text: str = "") -> None:
        normalized = normalize_invoice_link(url)
        if not normalized:
            return
        key = (normalized, anchor_text.lower().strip())
        if key in seen:
            return
        seen.add(key)
        seen_urls.add(normalized)
        candidates.append({"url": normalized, "anchor_text": anchor_text.strip()})

    parser = _AnchorLinkParser()
    try:
        parser.feed(html_text or "")
    except Exception:
        parser.links = []
    for href, anchor_text in parser.links:
        add(href, anchor_text)
    for text in (plain, html_text):
        for url in extract_urls(text):
            normalized = normalize_invoice_link(url)
            if normalized and normalized not in seen_urls:
                add(normalized, "")
    return candidates


def detect_link_platform(url: str, sender: str = "", subject: str = "") -> str:
    parsed = urlparse(url)
    host = (parsed.hostname or "").lower()
    combined = f"{url} {sender} {subject}".lower()
    if is_bwjf_tracking_link(url):
        return "埋点链接"
    if any(token in host for token in ("nfp.jss.com.cn", "jss.com.cn", "nuonuo.com")):
        return "诺诺网/JSS"
    if is_bwjf_invoice_link(url):
        return "云票/百望"
    if any(token in host for token in ("baiwang", "efapiao.com")):
        return "百望/票通"
    if "chinatax.gov.cn" in host:
        return "税务平台直链"
    if "fdfinvoice.com" in host:
        return "Apple硬件发票"
    if "apple.com" in host:
        return "Apple"
    if "jdcloud-oss.com" in host or "jd.com" in host:
        return "京东"
    if "bytedance.com" in host or "feishu" in combined or "larksuite" in combined:
        return "飞书"
    if "fin-invoice" in host or "taobao" in combined or "淘宝" in combined:
        return "淘宝闪购"
    if any(token in host for token in ("didi.com", "didichuxing.com", "didiglobal.com", "xiaojukeji.com", "udache.com", "diditaxi.com.cn")):
        return "滴滴"
    if any(token in host for token in ("caocaokeji.cn", "caocaomobility.com")) or "曹操出行" in combined:
        return "曹操出行"
    if "t3go.cn" in host or "t3出行" in combined:
        return "T3出行"
    if any(token in host for token in ("01zhuanche.com", "首汽约车")):
        return "首汽约车"
    if any(token in host for token in ("amap.com", "autonavi.com")) or "高德打车" in combined:
        return "高德打车"
    if "meituan.com" in host:
        return "美团"
    if any(token in host for token in ("trip.com", "ctrip.com")):
        return "携程/Trip"
    if any(token in host for token in ("12306", "railway")):
        return "12306"
    return "未知平台"


def link_has_invoice_signal(candidate: dict[str, str], subject: str, sender: str, body_text: str) -> bool:
    url = candidate.get("url", "")
    anchor_text = candidate.get("anchor_text", "")
    link_text = f"{url} {anchor_text}".lower()
    context_text = f"{subject} {sender} {body_text[:1000]}".lower()
    parsed = urlparse(url)
    ext = Path(parsed.path).suffix.lower()
    platform = detect_link_platform(url, sender, subject)
    if is_bwjf_tracking_link(url):
        return False
    if ext in INVOICE_EXTENSIONS or ext in ARCHIVE_EXTENSIONS:
        return True
    has_link_hint = any(hint.lower() in link_text for hint in LINK_INVOICE_HINTS)
    if platform in {"诺诺网/JSS", "云票/百望", "百望/票通", "税务平台直链", "Apple硬件发票", "京东", "飞书", "淘宝闪购", "滴滴", "曹操出行", "T3出行", "首汽约车", "高德打车"}:
        return True
    if platform != "未知平台" and has_link_hint:
        return True
    return has_link_hint


def is_noise_link(candidate: dict[str, str], subject: str, sender: str, body_text: str) -> bool:
    url = candidate.get("url", "")
    anchor_text = candidate.get("anchor_text", "")
    parsed = urlparse(url)
    host = (parsed.hostname or "").lower()
    path = (parsed.path or "").lower()
    combined = f"{url} {anchor_text}".lower()
    mail_context = f"{subject} {sender} {body_text}".lower()
    if is_bwjf_tracking_link(url):
        return True
    if any(hint in combined for hint in LINK_NOISE_HINTS):
        return True
    if host in {"ns.adobe.com", "purl.org"}:
        return True
    if host in {"wx.mail.qq.com", "support.apple.com", "www.apple.com", "apple.com"} and any(part in path for part in ("/legal", "/privacy", "/support")):
        return True
    if host in {"account.apple.com", "apps.apple.com"} and not any(token in combined for token in ("invoice", "发票", "receipt", "账单")):
        return True
    if link_has_invoice_signal(candidate, subject, sender, body_text):
        return False
    return "发票" not in mail_context and "invoice" not in mail_context


def select_invoice_link_candidates(
    candidates: list[dict[str, str]],
    subject: str,
    sender: str,
    body_text: str,
) -> list[dict[str, str]]:
    selected: list[dict[str, str]] = []
    seen_urls: set[str] = set()
    for candidate in candidates:
        url = candidate.get("url", "")
        if not url or url in seen_urls:
            continue
        if is_noise_link(candidate, subject, sender, body_text):
            continue
        if not link_has_invoice_signal(candidate, subject, sender, body_text):
            continue
        seen_urls.add(url)
        enriched = dict(candidate)
        enriched["platform"] = detect_link_platform(url, sender, subject)
        selected.append(enriched)
    selected.sort(key=lambda item: link_candidate_priority(item))
    return selected[:MAX_LINK_CANDIDATES_PER_MAIL]


def link_candidate_priority(candidate: dict[str, str]) -> tuple[int, str]:
    url = candidate.get("url", "").lower()
    anchor = candidate.get("anchor_text", "").lower()
    ext = Path(urlparse(url).path).suffix.lower()
    if is_bwjf_tracking_link(url):
        return (99, url)
    if preferred_bwjf_pdf_url(url) or is_bwjf_pdf_download_url(url):
        return (0, url)
    if ext == ".pdf":
        return (0, url)
    if "fdfinvoice.com" in url:
        return (0, url)
    if any(token in url for token in ("downloadpdf", "wjgs=pdf", "jflx=pdf")):
        return (1, url)
    if ext in {".ofd", ".xml", ".zip"} or any(token in url for token in ("downloadofd", "downloadxml", "wjgs=ofd", "wjgs=xml", "jflx=ofd", "jflx=xml")):
        return (4, url)
    if any(token in f"{url} {anchor}" for token in ("下载", "download", "发票", "invoice")):
        return (2, url)
    return (3, url)


def make_link_headers(env: dict[str, str]) -> dict[str, str]:
    return {
        "User-Agent": env.get("LINK_USER_AGENT") or env.get("MIMO_USER_AGENT") or "InvoiceCollector/1.0 (Local; Python)",
        "Accept": "application/pdf,application/xml,application/ofd,application/octet-stream,text/html;q=0.9,*/*;q=0.8",
    }


def infer_download_extension(url: str, content_type: str, content_disposition: str, data: bytes) -> str:
    combined = f"{url} {content_type} {content_disposition}".lower()
    stripped = data.lstrip()[:200].lower()
    if data.startswith(b"%PDF"):
        return ".pdf"
    if data.startswith(b"PK\x03\x04"):
        return ".zip"
    if "text/html" in content_type.lower() or stripped.startswith((b"<html", b"<!doctype html")):
        return ""
    if "application/zip" in combined or ".zip" in combined:
        return ".zip"
    if "application/pdf" in combined or ".pdf" in combined or "wjgs=pdf" in combined or "jflx=pdf" in combined or "pdfurl=" in combined or "pdfurl%3d" in combined:
        return ".pdf"
    if "ofd" in combined or ".ofd" in combined or "wjgs=ofd" in combined or "jflx=ofd" in combined:
        return ".ofd"
    if data.lstrip().startswith(b"<?xml") or "xml" in combined or ".xml" in combined or "wjgs=xml" in combined:
        return ".xml"
    return ""


def read_limited_response(response: object, limit: int = MAX_LINK_DOWNLOAD_BYTES) -> bytes:
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = response.read(1024 * 512)  # type: ignore[attr-defined]
        if not chunk:
            break
        total += len(chunk)
        if total > limit:
            raise RuntimeError("download_too_large")
        chunks.append(chunk)
    return b"".join(chunks)


def fetch_url_bytes(url: str, env: dict[str, str], retry_attempts: int, retry_sleep: float) -> tuple[bytes, str, str, str]:
    headers = make_link_headers(env)
    if "chinatax.gov.cn" in (urlparse(url).hostname or "").lower():
        headers["User-Agent"] = env.get("LINK_USER_AGENT") or "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"

    def urllib_once() -> tuple[bytes, str, str, str]:
        req = urlrequest.Request(url, headers=headers)
        with urlrequest.urlopen(req, timeout=LINK_DOWNLOAD_TIMEOUT_SECONDS) as response:
            data = read_limited_response(response)
            content_type = str(response.headers.get("Content-Type") or "")
            disposition = str(response.headers.get("Content-Disposition") or "")
            resolved_url = str(response.geturl() or url)
        return data, content_type, disposition, resolved_url

    def once() -> tuple[bytes, str, str, str]:
        import httpx

        try:
            with httpx.Client(
                follow_redirects=True,
                headers=headers,
                timeout=LINK_DOWNLOAD_TIMEOUT_SECONDS,
                trust_env=False,
            ) as client:
                response = client.get(url)
                response.raise_for_status()
                data = response.content
        except Exception:
            if "chinatax.gov.cn" in (urlparse(url).hostname or "").lower():
                return urllib_once()
            raise
        if len(data) > MAX_LINK_DOWNLOAD_BYTES:
            raise RuntimeError("download_too_large")
        content_type = str(response.headers.get("Content-Type") or "")
        disposition = str(response.headers.get("Content-Disposition") or "")
        resolved_url = str(response.url or url)
        return data, content_type, disposition, resolved_url

    return retry_call(once, attempts=retry_attempts, sleep_seconds=retry_sleep, label=f"link download {urlparse(url).netloc}")


def resolve_bwjf_pdf_url(url: str, env: dict[str, str], retry_attempts: int, retry_sleep: float) -> str:
    if is_bwjf_tracking_link(url):
        return ""
    direct = preferred_bwjf_pdf_url(url)
    if direct:
        return direct
    if is_bwjf_pdf_download_url(url):
        return url

    parsed = urlparse(url)
    host = (parsed.hostname or "").lower()
    path = (parsed.path or "").lower()
    if host != "fp.bwjf.cn" or not path.startswith("/u/"):
        return ""

    headers = make_link_headers(env)

    def once() -> str:
        import httpx

        with httpx.Client(
            follow_redirects=False,
            headers=headers,
            timeout=LINK_DOWNLOAD_TIMEOUT_SECONDS,
            trust_env=False,
        ) as client:
            response = client.get(url)
        location = str(response.headers.get("Location") or "").strip()
        if location:
            resolved = urljoin(url, location)
            nested = preferred_bwjf_pdf_url(resolved)
            if nested:
                return nested
            if is_bwjf_pdf_download_url(resolved):
                return resolved

        with httpx.Client(
            follow_redirects=True,
            headers=headers,
            timeout=LINK_DOWNLOAD_TIMEOUT_SECONDS,
            trust_env=False,
        ) as client:
            response = client.get(url)
            response.raise_for_status()
        final_url = str(response.url or url)
        nested = preferred_bwjf_pdf_url(final_url)
        if nested:
            return nested
        text = response.text[:200000]
        for candidate in html_candidate_links(text.encode("utf-8", errors="replace"), final_url):
            candidate_url = candidate.get("url", "")
            nested = preferred_bwjf_pdf_url(candidate_url)
            if nested:
                return nested
            if is_bwjf_pdf_download_url(candidate_url):
                return candidate_url
        return ""

    return retry_call(once, attempts=retry_attempts, sleep_seconds=retry_sleep, label=f"BWJF PDF resolve {urlparse(url).netloc}")


def fetch_bwjf_pdf_artifact(
    url: str,
    env: dict[str, str],
    target_dir: Path,
    prefix: str,
    retry_attempts: int,
    retry_sleep: float,
) -> list[dict[str, str]]:
    pdf_url = resolve_bwjf_pdf_url(url, env, retry_attempts, retry_sleep)
    if not pdf_url:
        return []
    data, content_type, disposition, resolved_url = fetch_url_bytes(pdf_url, env, retry_attempts, retry_sleep)
    ext = infer_download_extension(resolved_url, content_type, disposition, data)
    if ext != ".pdf":
        return []
    saved, filename = save_response_artifact(resolved_url, disposition, data, target_dir, prefix, ".pdf")
    return [
        {
            "path": str(saved),
            "original_name": filename,
            "source_url": url,
            "resolved_url": resolved_url,
            "platform": "云票/百望",
            "status": "Link_Downloaded",
        }
    ]


def is_nuonuo_short_invoice_link(url: str) -> bool:
    parsed = urlparse(url)
    host = (parsed.hostname or "").lower()
    if "nnfp.jss.com.cn" not in host:
        return False
    path = (parsed.path or "").strip("/")
    return bool(path) and not path.startswith(("allow/", "scan-invoice/", "invoice/", "scan2/"))


def fetch_nuonuo_pdf_artifact(
    url: str,
    env: dict[str, str],
    retry_attempts: int,
    retry_sleep: float,
) -> tuple[bytes, str, str, str] | None:
    headers = make_link_headers(env)
    headers.setdefault("Referer", url)

    def once() -> tuple[bytes, str, str, str] | None:
        import httpx

        with httpx.Client(follow_redirects=True, headers=headers, timeout=LINK_DOWNLOAD_TIMEOUT_SECONDS, trust_env=False) as client:
            landing = client.get(url)
            landing.raise_for_status()
            resolved_url = str(landing.url or url)
            query = dict(parse_qsl(urlparse(resolved_url).query, keep_blank_values=True))
            param_list = query.get("paramList", "")
            if not param_list:
                return None
            payload = {
                "paramList": param_list,
                "code": query.get("code", ""),
                "aliView": query.get("aliView", ""),
                "invoiceDetailMiddleUri": "",
                "shortLinkSource": query.get("shortLinkSource", ""),
            }
            api = client.post(
                "https://nnfp.jss.com.cn/scan2/getIvcDetailShow.do",
                data=payload,
                headers={
                    "Content-Type": "application/x-www-form-urlencoded",
                    "User-Agent": headers["User-Agent"],
                    "Referer": resolved_url,
                },
            )
            api.raise_for_status()
            body = api.json()
            if str(body.get("status")) != "0000":
                return None
            invoice = ((body.get("data") or {}).get("invoiceSimpleVo") or {})
            pdf_url = invoice.get("url") or ""
            if not pdf_url:
                return None
            pdf = client.get(str(pdf_url), headers={"User-Agent": headers["User-Agent"], "Referer": resolved_url})
            pdf.raise_for_status()
            content_type = str(pdf.headers.get("Content-Type") or "application/pdf")
            disposition = str(pdf.headers.get("Content-Disposition") or "")
            return pdf.content, content_type, disposition, str(pdf.url or pdf_url)

    return retry_call(once, attempts=retry_attempts, sleep_seconds=retry_sleep, label=f"Nuonuo invoice API {urlparse(url).netloc}")


def save_downloaded_artifact(data: bytes, target_dir: Path, filename: str, prefix: str, ext: str) -> Path:
    target_dir.mkdir(parents=True, exist_ok=True)
    safe_filename = safe_name(filename, f"{prefix}{ext}")
    if not Path(safe_filename).suffix and ext:
        safe_filename = f"{safe_filename}{ext}"
    if ext and Path(safe_filename).suffix.lower() != ext:
        safe_filename = f"{Path(safe_filename).stem}{ext}"
    target = target_dir / f"{prefix}_{safe_filename}"
    counter = 2
    while target.exists():
        target = target_dir / f"{prefix}_{counter}_{safe_filename}"
        counter += 1
    target.write_bytes(data)
    return target


def html_candidate_links(data: bytes, base_url: str) -> list[dict[str, str]]:
    text = data.decode("utf-8", errors="replace")
    parser = _AnchorLinkParser()
    try:
        parser.feed(text)
    except Exception:
        parser.links = []
    candidates = [{"url": href, "anchor_text": anchor_text} for href, anchor_text in parser.links]
    for url in extract_urls(text):
        candidates.append({"url": url, "anchor_text": ""})
    normalized: list[dict[str, str]] = []
    seen: set[str] = set()
    for candidate in candidates:
        url = normalize_invoice_link(candidate.get("url", ""), base_url=base_url)
        if not url or url in seen:
            continue
        seen.add(url)
        normalized.append({"url": url, "anchor_text": candidate.get("anchor_text", "")})
    return normalized


def save_response_artifact(
    response_url: str,
    content_disposition: str,
    data: bytes,
    target_dir: Path,
    prefix: str,
    ext: str,
) -> tuple[Path, str]:
    match = re.search(r"filename\*?=(?:UTF-8''|\"?)([^\";]+)", content_disposition, flags=re.IGNORECASE)
    if match:
        filename = safe_name(unquote(match.group(1)))
    else:
        path_name = Path(urlparse(response_url).path).name
        filename = safe_name(path_name, f"invoice_link{ext}") if path_name else f"invoice_link{ext}"
    saved = save_downloaded_artifact(data, target_dir, filename, prefix, ext)
    return saved, filename


def extract_zip_invoice_artifacts(
    data: bytes,
    target_dir: Path,
    prefix: str,
    source_url: str,
    resolved_url: str,
    platform: str,
) -> list[dict[str, str]]:
    artifacts: list[dict[str, str]] = []
    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile:
        return artifacts

    with archive:
        for index, info in enumerate(archive.infolist(), start=1):
            if info.is_dir():
                continue
            ext = Path(info.filename).suffix.lower()
            if ext not in INVOICE_EXTENSIONS:
                continue
            payload = archive.read(info)
            if ext == ".pdf" and not payload.startswith(b"%PDF"):
                continue
            if ext == ".xml":
                stripped = payload.lstrip()
                if not stripped.startswith(b"<?xml"):
                    continue
                head = stripped[:4096].lower()
                if b"<einvoice" not in head and b"<invoice" not in head:
                    continue
            filename = safe_name(Path(info.filename).name, f"invoice_zip_{index}{ext}")
            saved = save_downloaded_artifact(payload, target_dir, filename, f"{prefix}_zip{index}", ext)
            artifacts.append(
                {
                    "path": str(saved),
                    "original_name": filename,
                    "source_url": source_url,
                    "resolved_url": resolved_url,
                    "platform": platform,
                    "status": "Link_Downloaded",
                }
            )
    return artifacts


def xml_local_name(tag: str) -> str:
    return str(tag or "").rsplit("}", 1)[-1]


def first_xml_text(root: ET.Element, *names: str) -> str:
    wanted = {name.lower() for name in names}
    for elem in root.iter():
        if xml_local_name(elem.tag).lower() in wanted and elem.text and elem.text.strip():
            return elem.text.strip()
    return ""


def nested_xml_text(root: ET.Element, parent_name: str, child_name: str) -> str:
    for parent in root.iter():
        if xml_local_name(parent.tag).lower() != parent_name.lower():
            continue
        for child in parent.iter():
            if xml_local_name(child.tag).lower() == child_name.lower() and child.text and child.text.strip():
                return child.text.strip()
    return ""


def decode_invoice_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def normalize_structured_invoice_kind(value: str) -> str:
    text = str(value or "")
    if "专用" in text:
        return "增值税专用发票"
    if "普通" in text or "电子发票" in text:
        return "增值税普通发票"
    return text.strip()


def categorize_structured_invoice(*texts: str) -> str:
    combined = "".join(texts)
    if any(token in combined for token in ("餐饮", "餐费", "饭店", "餐厅", "酒楼", "美食")):
        return "餐饮"
    if any(token in combined for token in ("物流", "配送", "闪购")):
        return "平台服务"
    if any(token in combined for token in ("服装", "衣服")):
        return "服装"
    if any(token in combined for token in ("电脑", "手机", "平板", "电子计算机")):
        return "电子设备"
    return "其他"


def parse_invoice_xml_bytes(data: bytes) -> tuple[dict[str, str], str]:
    text = decode_invoice_text(data)
    try:
        root = ET.fromstring(text.encode("utf-8"))
    except Exception:
        fields = parse_invoice_fields_from_text(text)
        return fields, text

    fields: dict[str, str] = {}
    fields["invoice_number"] = first_xml_text(root, "InvoiceNumber", "EIid", "InvoiceNo", "Fphm")
    issue_time = first_xml_text(root, "IssueTime", "RequestTime", "Kprq", "InvoiceDate")
    date_match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", issue_time)
    if date_match:
        fields["invoice_date"] = f"{date_match.group(1)}-{int(date_match.group(2)):02d}-{int(date_match.group(3)):02d}"
    fields["seller"] = first_xml_text(root, "SellerName", "Xsfmc")
    fields["purchaser"] = first_xml_text(root, "BuyerName", "PurchaserName", "Gmfmc")
    amount = first_xml_text(root, "TotalTax-includedAmount", "TotalTaxIncludedAmount", "TotalAmount", "Jshj")
    if amount:
        amount_match = re.search(r"([0-9,]+(?:\.[0-9]{1,2})?)", amount)
        if amount_match:
            fields["amount"] = f"{float(amount_match.group(1).replace(',', '')):.2f}"
    general_type = nested_xml_text(root, "GeneralOrSpecialVAT", "LabelName")
    invoice_type = nested_xml_text(root, "EInvoiceType", "LabelName")
    fields["invoice_kind"] = normalize_structured_invoice_kind(general_type or invoice_type)
    item_name = first_xml_text(root, "ItemName", "GoodsName", "Xmmc")
    fields["category"] = categorize_structured_invoice(fields.get("seller", ""), item_name)

    fallback = parse_invoice_fields_from_text(text)
    for key, value in fallback.items():
        fields.setdefault(key, value)
    return {key: value for key, value in fields.items() if value}, text


def parse_ofd_invoice_bytes(data: bytes) -> tuple[dict[str, str], str]:
    texts: list[str] = []
    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile:
        return {}, ""
    best_fields: dict[str, str] = {}
    with archive:
        for info in archive.infolist():
            if info.is_dir() or not info.filename.lower().endswith(".xml") or info.file_size > 1024 * 1024:
                continue
            payload = archive.read(info)
            fields, text = parse_invoice_xml_bytes(payload)
            if text:
                texts.append(text)
            if has_minimum_invoice_fields(fields) and fields.get("invoice_number"):
                return fields, text
            if len(fields) > len(best_fields):
                best_fields = fields
    return best_fields, "\n".join(texts)


def parse_structured_invoice_artifact(path: Path) -> dict[str, object]:
    suffix = path.suffix.lower()
    data = path.read_bytes()
    if suffix == ".xml":
        fields, raw_text = parse_invoice_xml_bytes(data)
        engine = "xml_structured"
    elif suffix == ".ofd":
        fields, raw_text = parse_ofd_invoice_bytes(data)
        engine = "ofd_structured"
    else:
        return {}
    return {
        "status": "parsed" if has_minimum_invoice_fields(fields) else "need_review",
        "engine": engine,
        "page_number": 1,
        "page_count": 1,
        "raw_text": raw_text[:20000],
        "error": "" if has_minimum_invoice_fields(fields) else "structured_invoice_incomplete",
        **fields,
    }


def fetch_apple_fdfinvoice_artifacts(
    url: str,
    env: dict[str, str],
    target_dir: Path,
    prefix: str,
    retry_attempts: int,
    retry_sleep: float,
) -> list[dict[str, str]]:
    parsed = urlparse(url)
    order_ref_no = dict(parse_qsl(parsed.query, keep_blank_values=True)).get("orderRefNo", "")
    if not order_ref_no:
        return []

    endpoint = "https://www.fdfinvoice.com/prod-api/output/unauth/getDeliverDataByOrderRefNo"
    headers = make_link_headers(env)
    headers.update(
        {
            "Accept": "application/json, text/plain, */*",
            "Content-Type": "application/json;charset=UTF-8",
            "Origin": "https://www.fdfinvoice.com",
            "Referer": url,
        }
    )

    def once() -> list[dict[str, str]]:
        req = urlrequest.Request(endpoint, data=order_ref_no.encode("utf-8"), headers=headers, method="POST")
        with urlrequest.urlopen(req, timeout=LINK_DOWNLOAD_TIMEOUT_SECONDS) as response:
            body = response.read(MAX_LINK_DOWNLOAD_BYTES)
        payload = json.loads(body.decode("utf-8", errors="replace"))
        if str(payload.get("code")) != "200":
            raise RuntimeError(str(payload.get("msg") or "apple_fdfinvoice_failed"))
        result: list[dict[str, str]] = []
        for item in payload.get("data") or []:
            try:
                amount = float(str(item.get("totalTaxAmount") or "0").replace(",", ""))
            except ValueError:
                amount = 0.0
            buyer = str(item.get("buyerName") or "")
            if item.get("expireFlag") or str(item.get("isRed") or "") == "1" or amount <= 0:
                continue
            if buyer in {"个人", "个人用户"}:
                continue
            pdf_url = str(item.get("pdfUrl") or "")
            if not pdf_url:
                continue
            pdf_data, content_type, disposition, resolved_url = fetch_url_bytes(pdf_url, env, retry_attempts, retry_sleep)
            ext = infer_download_extension(resolved_url, content_type, disposition, pdf_data) or ".pdf"
            if ext != ".pdf":
                continue
            saved, filename = save_response_artifact(resolved_url, disposition, pdf_data, target_dir, prefix, ext)
            result.append(
                {
                    "path": str(saved),
                    "original_name": filename,
                    "source_url": url,
                    "resolved_url": resolved_url,
                    "platform": "Apple硬件发票",
                    "status": "Link_Downloaded",
                }
            )
        return result

    return retry_call(once, attempts=retry_attempts, sleep_seconds=retry_sleep, label="Apple hardware invoice download")


def download_invoice_links(
    candidates: list[dict[str, str]],
    env: dict[str, str],
    target_dir: Path,
    prefix: str,
    retry_attempts: int,
    retry_sleep: float,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    artifacts: list[dict[str, str]] = []
    unresolved: list[dict[str, str]] = []
    queued = list(candidates)
    seen: set[str] = set()

    while queued and len(seen) < MAX_LINK_CANDIDATES_PER_MAIL * 2:
        candidate = queued.pop(0)
        url = candidate.get("url", "")
        if not url or url in seen:
            continue
        seen.add(url)
        platform = candidate.get("platform") or detect_link_platform(url)
        try:
            if platform == "Apple硬件发票":
                apple_artifacts = fetch_apple_fdfinvoice_artifacts(
                    url,
                    env,
                    target_dir,
                    prefix,
                    retry_attempts,
                    retry_sleep,
                )
                if apple_artifacts:
                    artifacts.extend(apple_artifacts)
                    continue

            if platform == "云票/百望" or is_bwjf_invoice_link(url):
                bwjf_artifacts = fetch_bwjf_pdf_artifact(
                    url,
                    env,
                    target_dir,
                    prefix,
                    retry_attempts,
                    retry_sleep,
                )
                if bwjf_artifacts:
                    artifacts.extend(bwjf_artifacts)
                    continue

            if is_nuonuo_short_invoice_link(url):
                nuonuo_result = fetch_nuonuo_pdf_artifact(url, env, retry_attempts, retry_sleep)
                if nuonuo_result:
                    data, content_type, disposition, resolved_url = nuonuo_result
                    ext = infer_download_extension(resolved_url, content_type, disposition, data) or ".pdf"
                    saved, filename = save_response_artifact(resolved_url, disposition, data, target_dir, prefix, ext)
                    artifacts.append(
                        {
                            "path": str(saved),
                            "original_name": filename,
                            "source_url": url,
                            "resolved_url": resolved_url,
                            "platform": platform,
                            "status": "Link_Downloaded",
                        }
                    )
                    continue

            data, content_type, disposition, resolved_url = fetch_url_bytes(url, env, retry_attempts, retry_sleep)
            ext = infer_download_extension(resolved_url, content_type, disposition, data)
            if ext:
                if ext == ".zip":
                    zip_artifacts = extract_zip_invoice_artifacts(data, target_dir, prefix, url, resolved_url, platform)
                    if zip_artifacts:
                        artifacts.extend(zip_artifacts)
                        continue
                    unresolved.append(
                        {
                            "url": url,
                            "platform": platform,
                            "status": "Non_Invoice_Link",
                            "note": "链接下载结果为 ZIP，但压缩包内未发现 PDF/OFD/XML 发票文件。",
                        }
                    )
                    continue
                saved, filename = save_response_artifact(resolved_url, disposition, data, target_dir, prefix, ext)
                artifacts.append(
                    {
                        "path": str(saved),
                        "original_name": filename,
                        "source_url": url,
                        "resolved_url": resolved_url,
                        "platform": platform,
                        "status": "Link_Downloaded",
                    }
                )
                continue

            if "html" in content_type.lower() or b"<html" in data[:500].lower():
                nested = select_invoice_link_candidates(
                    html_candidate_links(data, resolved_url),
                    "",
                    platform,
                    data.decode("utf-8", errors="replace")[:5000],
                )
                for nested_candidate in nested:
                    nested_candidate.setdefault("platform", platform)
                    if nested_candidate.get("url") not in seen:
                        queued.append(nested_candidate)
                if nested:
                    continue

            if content_type.lower().startswith("image/"):
                unresolved.append(
                    {
                        "url": url,
                        "platform": platform,
                        "status": "Non_Invoice_Link",
                        "note": f"链接下载结果为图片，不是可解析发票文件；Content-Type={content_type[:80]}",
                    }
                )
                continue

            unresolved.append(
                {
                    "url": url,
                    "platform": platform,
                    "status": "Link_Need_Manual",
                    "note": f"链接可访问但没有发现 PDF/OFD/XML 下载结果；Content-Type={content_type[:80]}",
                }
            )
        except Exception as exc:
            unresolved.append(
                {
                    "url": url,
                    "platform": platform,
                    "status": "Link_Need_Manual",
                    "note": f"自动下载失败，需要人工打开确认：{str(exc)[:180]}",
                }
            )

    return artifacts, unresolved


def html_image_sources(html_text: str) -> list[dict[str, str]]:
    parser = _ImageSourceParser()
    try:
        parser.feed(html_text or "")
    except Exception:
        return []
    return parser.images


def message_content_id_map(msg: Message) -> dict[str, Message]:
    result: dict[str, Message] = {}
    for part in msg.walk():
        content_id = str(part.get("Content-ID") or "").strip().strip("<>")
        if content_id:
            result[content_id] = part
    return result


def decode_data_url_image(src: str) -> tuple[bytes, str] | None:
    match = re.match(r"data:(image/[a-zA-Z0-9.+-]+);base64,(.+)", src, flags=re.DOTALL)
    if not match:
        return None
    try:
        return base64.b64decode(match.group(2), validate=False), match.group(1)
    except Exception:
        return None


def image_bytes_to_data_url(data: bytes, content_type: str) -> str:
    mime = content_type if content_type.startswith("image/") else "image/png"
    return f"data:{mime};base64,{base64.b64encode(data).decode('ascii')}"


def has_qr_context(*texts: str) -> bool:
    combined = " ".join(texts)
    return any(hint in combined for hint in QR_HINTS)


def is_decorative_image_name(name: str) -> bool:
    lower_name = name.lower()
    return any(hint in lower_name for hint in DECORATIVE_IMAGE_NAME_HINTS)


def add_qr_candidate(
    candidates: list[dict[str, object]],
    seen_hashes: set[str],
    *,
    data: bytes,
    content_type: str,
    filename: str,
    source: str,
) -> None:
    if not data or len(data) < 512:
        return
    digest = hashlib.sha256(data).hexdigest()
    if digest in seen_hashes:
        return
    seen_hashes.add(digest)
    candidates.append(
        {
            "data": data,
            "content_type": content_type or "image/png",
            "filename": safe_name(filename, "qr_clue.png"),
            "source": source,
        }
    )


def collect_qr_image_candidates(
    msg: Message,
    plain: str,
    html_text: str,
    env: dict[str, str],
    retry_attempts: int,
    retry_sleep: float,
) -> list[dict[str, object]]:
    qr_context = has_qr_context(plain, html_text)
    if not qr_context:
        return []

    candidates: list[dict[str, object]] = []
    seen_hashes: set[str] = set()
    cid_parts = message_content_id_map(msg)

    for image in html_image_sources(html_text):
        src = image.get("src", "").strip()
        hint_text = f"{src} {image.get('alt', '')} {image.get('title', '')}"
        if not has_qr_context(hint_text, plain, html_text):
            continue
        data_url = decode_data_url_image(src)
        if data_url:
            data, content_type = data_url
            add_qr_candidate(candidates, seen_hashes, data=data, content_type=content_type, filename="html_base64_qr.png", source="html_base64")
            continue
        if src.lower().startswith("cid:"):
            cid = src[4:].strip().strip("<>")
            part = cid_parts.get(cid)
            if part:
                payload = part.get_payload(decode=True) or b""
                add_qr_candidate(
                    candidates,
                    seen_hashes,
                    data=payload,
                    content_type=part.get_content_type() or "image/png",
                    filename=part_filename(part),
                    source=f"cid:{cid}",
                )
            continue
        normalized_src = normalize_invoice_link(src)
        if normalized_src and (has_qr_context(hint_text) or detect_link_platform(normalized_src) != "未知平台"):
            try:
                data, content_type, _disposition, resolved_url = fetch_url_bytes(normalized_src, env, retry_attempts, retry_sleep)
                if content_type.lower().startswith("image/"):
                    add_qr_candidate(
                        candidates,
                        seen_hashes,
                        data=data,
                        content_type=content_type,
                        filename=Path(urlparse(resolved_url).path).name or "remote_qr.png",
                        source=resolved_url,
                    )
            except Exception:
                continue

    for part in msg.walk():
        content_type = (part.get_content_type() or "").lower()
        if not content_type.startswith("image/"):
            continue
        filename = part_filename(part)
        if is_decorative_image_name(filename):
            continue
        payload = part.get_payload(decode=True) or b""
        add_qr_candidate(
            candidates,
            seen_hashes,
            data=payload,
            content_type=content_type,
            filename=filename,
            source="mail_image_part",
        )

    return candidates


def save_qr_clue_candidate(candidate: dict[str, object], prefix: str) -> Path:
    target_dir = MANUAL_DIR / "二维码线索"
    target_dir.mkdir(parents=True, exist_ok=True)
    filename = safe_name(str(candidate.get("filename") or "qr_clue.png"), "qr_clue.png")
    target = target_dir / f"{prefix}_{filename}"
    counter = 2
    while target.exists():
        target = target_dir / f"{prefix}_{counter}_{filename}"
        counter += 1
    target.write_bytes(candidate.get("data") if isinstance(candidate.get("data"), bytes) else b"")
    return target


def recover_qr_invoice_links(
    qr_candidates: list[dict[str, object]],
    qr_url_extract: Callable[[list[str]], str] | None,
    env: dict[str, str],
    target_dir: Path,
    prefix: str,
    retry_attempts: int,
    retry_sleep: float,
    subject: str,
    sender: str,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    artifacts: list[dict[str, str]] = []
    unresolved: list[dict[str, str]] = []
    if not qr_candidates:
        return artifacts, unresolved

    if not qr_url_extract:
        for candidate in qr_candidates:
            saved = save_qr_clue_candidate(candidate, prefix)
            unresolved.append(
                {
                    "url": str(saved),
                    "platform": "",
                    "status": "QR_Need_Manual",
                    "note": "未配置视觉二维码解析器，二维码图片已保存到人工复核。",
                }
            )
        return artifacts, unresolved

    for index, candidate in enumerate(qr_candidates, start=1):
        data = candidate.get("data")
        if not isinstance(data, bytes):
            continue
        content_type = str(candidate.get("content_type") or "image/png")
        try:
            qr_url = qr_url_extract([image_bytes_to_data_url(data, content_type)]).strip()
        except Exception as exc:
            qr_url = ""
            error_note = f"视觉模型识别二维码失败：{str(exc)[:180]}"
        else:
            error_note = "视觉模型未能识别出二维码 URL。"

        qr_url = normalize_invoice_link(qr_url)
        if not qr_url:
            saved = save_qr_clue_candidate(candidate, f"{prefix}_{index}")
            unresolved.append(
                {
                    "url": str(saved),
                    "platform": "",
                    "status": "QR_Need_Manual",
                    "note": error_note,
                }
            )
            continue

        link_candidate = {
            "url": qr_url,
            "anchor_text": "二维码识别",
            "platform": detect_link_platform(qr_url, sender, subject),
        }
        downloaded, failed = download_invoice_links(
            [link_candidate],
            env,
            target_dir,
            f"{prefix}_{index}",
            retry_attempts,
            retry_sleep,
        )
        if downloaded:
            artifacts.extend(downloaded)
            continue

        saved = save_qr_clue_candidate(candidate, f"{prefix}_{index}")
        note = failed[0].get("note", "二维码 URL 下载失败，需要人工确认。") if failed else "二维码 URL 下载失败，需要人工确认。"
        unresolved.append(
            {
                "url": qr_url,
                "platform": link_candidate.get("platform", ""),
                "status": "QR_Need_Manual",
                "note": f"{note} 已保存二维码图片：{saved}",
            }
        )

    return artifacts, unresolved


def maybe_save_qr_clues(msg: Message, target_dir: Path, prefix: str, plain: str, html_text: str) -> list[Path]:
    qr_hint = any(token in f"{plain} {html_text}" for token in ("二维码", "扫码", "扫一扫", "QR", "qr"))
    if not qr_hint:
        return []
    target_dir = MANUAL_DIR / "二维码线索"
    saved: list[Path] = []
    for part in msg.walk():
        content_type = (part.get_content_type() or "").lower()
        filename = part_filename(part)
        ext = Path(filename).suffix.lower()
        if not (content_type.startswith("image/") or ext in SKIPPED_IMAGE_EXTENSIONS):
            continue
        if any(hint in filename.lower() for hint in DECORATIVE_IMAGE_NAME_HINTS):
            continue
        payload = part.get_payload(decode=True)
        if not payload or len(payload) < MIN_IMAGE_SIGNATURE_BYTES:
            continue
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / f"{prefix}_{safe_name(filename, 'qr_clue')}"
        counter = 2
        while target.exists():
            target = target_dir / f"{prefix}_{counter}_{safe_name(filename, 'qr_clue')}"
            counter += 1
        target.write_bytes(payload)
        saved.append(target)
    return saved


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def save_attachment(part: Message, target_dir: Path, prefix: str) -> tuple[Path, str] | None:
    filename = part_filename(part)
    if not should_download_attachment(part, filename):
        return None
    payload = part.get_payload(decode=True)
    if not payload:
        return None
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"{prefix}_{filename}"
    counter = 2
    while target.exists():
        target = target_dir / f"{prefix}_{counter}_{filename}"
        counter += 1
    target.write_bytes(payload)
    return target, filename


class ProcessLog:
    def __init__(self, db_path: Path):
        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(str(self.db_path))
        self.conn.execute(
            """
            CREATE TABLE IF NOT EXISTS processed_messages (
                mailbox TEXT NOT NULL,
                uid TEXT NOT NULL,
                status TEXT NOT NULL,
                records_count INTEGER NOT NULL DEFAULT 0,
                processed_at TEXT NOT NULL,
                error TEXT NOT NULL DEFAULT '',
                PRIMARY KEY (mailbox, uid)
            )
            """
        )
        self.conn.commit()

    def is_processed(self, mailbox: str, uid: bytes | str) -> bool:
        uid_text = uid.decode("ascii", errors="replace") if isinstance(uid, bytes) else str(uid)
        row = self.conn.execute(
            "SELECT status FROM processed_messages WHERE mailbox = ? AND uid = ?",
            (mailbox, uid_text),
        ).fetchone()
        if row is None:
            return False
        return str(row[0] or "") not in {"failed", "fetch_failed", "timeout", "imap_timeout"}

    def mark(self, mailbox: str, uid: bytes | str, status: str, records_count: int = 0, error: str = "") -> None:
        uid_text = uid.decode("ascii", errors="replace") if isinstance(uid, bytes) else str(uid)
        self.conn.execute(
            """
            INSERT INTO processed_messages (mailbox, uid, status, records_count, processed_at, error)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(mailbox, uid) DO UPDATE SET
                status = excluded.status,
                records_count = excluded.records_count,
                processed_at = excluded.processed_at,
                error = excluded.error
            """,
            (mailbox, uid_text, status, int(records_count), dt.datetime.now().isoformat(timespec="seconds"), error[:500]),
        )
        self.conn.commit()

    def close(self) -> None:
        self.conn.close()


def parse_internal_date(metadata: object, fallback: str) -> str:
    if isinstance(metadata, bytes):
        parsed = imaplib.Internaldate2tuple(metadata)
        if parsed:
            return dt.datetime.fromtimestamp(time.mktime(parsed)).strftime("%Y-%m-%d %H:%M:%S")
    return fallback


def uid_sort_key(uid: bytes | str) -> int:
    uid_text = uid.decode("ascii", errors="ignore") if isinstance(uid, bytes) else str(uid)
    try:
        return int(uid_text)
    except ValueError:
        return 0


def imap_quote(value: str) -> str:
    return '"' + value.replace("\\", "\\\\").replace('"', '\\"') + '"'


def uid_search(
    mail: imaplib.IMAP4_SSL,
    criterion: str | list[str | bytes],
    *,
    retry_attempts: int,
    retry_sleep: float,
    label: str,
    use_utf8: bool = False,
) -> set[bytes]:
    if isinstance(criterion, list):
        args: tuple[object, ...] = (None, *criterion)
    else:
        args = ("CHARSET", "UTF-8", criterion.encode("utf-8")) if use_utf8 else (None, criterion)
    typ, data = retry_call(
        lambda: mail.uid("SEARCH", *args),
        attempts=retry_attempts,
        sleep_seconds=retry_sleep,
        label=label,
    )
    if typ != "OK":
        raise RuntimeError(f"{label} failed: {typ}")
    return set(data[0].split()) if data and data[0] else set()


def bodystructure_has_invoice_attachment(raw_parts: list[object]) -> bool:
    chunks: list[bytes] = []
    for item in raw_parts:
        if isinstance(item, tuple):
            chunks.extend(part for part in item if isinstance(part, bytes))
        elif isinstance(item, bytes):
            chunks.append(item)
    text = b" ".join(chunks).decode("utf-8", errors="ignore")
    decoded_text = decode_mime_words(text)
    haystack = f"{text}\n{decoded_text}".lower()
    return any(
        token in haystack
        for token in (
            "application/pdf",
            ".pdf",
            ".ofd",
            ".xml",
            "application/vnd.openxmlformats-officedocument",
        )
    )


def sender_is_priority(sender: str) -> bool:
    sender_lower = str(sender or "").lower()
    return any(domain.lower() in sender_lower for domain in PRIORITY_SENDER_DOMAINS)


def search_attachment_candidate_uids(
    mail: imaplib.IMAP4_SSL,
    uids: set[bytes],
    *,
    retry_attempts: int,
    retry_sleep: float,
    max_probe: int = MAX_ATTACHMENT_PROBE_UIDS,
) -> set[bytes]:
    result: set[bytes] = set()
    ordered = sorted(uids, key=uid_sort_key, reverse=True)
    if len(ordered) > max_probe:
        print(f"Sample Calibration attachment probe skipped: date_matches={len(ordered)} exceeds max_probe={max_probe}")
        return result
    for uid in ordered:
        try:
            typ, data = retry_call(
                lambda uid=uid: mail.uid("FETCH", uid, "(BODYSTRUCTURE)"),
                attempts=retry_attempts,
                sleep_seconds=retry_sleep,
                label=f"IMAP attachment probe UID {uid.decode('ascii', errors='replace')}",
            )
        except Exception:
            continue
        if typ == "OK" and data and bodystructure_has_invoice_attachment(list(data)):
            result.add(uid)
    return result


def search_invoice_candidate_uids(
    mail: imaplib.IMAP4_SSL,
    *,
    since: str,
    until: str,
    retry_attempts: int,
    retry_sleep: float,
    limit: int | None,
    enable_attachment_probe: bool = True,
    max_attachment_probe: int = MAX_ATTACHMENT_PROBE_UIDS,
    search_mode: str = "filtered",
) -> list[bytes]:
    since_atom = date_to_imap(since)
    before_atom = date_to_imap(next_date(until))
    date_uids = uid_search(
        mail,
        ["SINCE", since_atom, "BEFORE", before_atom],
        retry_attempts=retry_attempts,
        retry_sleep=retry_sleep,
        label="IMAP date search",
    )
    if search_mode.lower() in {"date_only", "date-only", "basic"}:
        ordered = sorted(date_uids, key=uid_sort_key, reverse=True)
        if limit:
            ordered = ordered[:limit]
        print(
            "Sample Calibration IMAP filter: "
            f"mode=date_only, date_matches={len(date_uids)}, candidates={len(ordered)}"
        )
        return ordered

    subject_uids: set[bytes] = set()
    priority_uids: set[bytes] = set()

    for keyword in SUBJECT_KEYWORDS:
        use_utf8 = any(ord(char) > 127 for char in keyword)
        try:
            subject_uids.update(
                uid_search(
                    mail,
                    f"(SUBJECT {imap_quote(keyword)})",
                    retry_attempts=retry_attempts,
                    retry_sleep=retry_sleep,
                    label=f"IMAP subject search {keyword}",
                    use_utf8=use_utf8,
                )
            )
        except Exception as exc:
            print(f"Sample Calibration filter warning: subject keyword '{keyword}' skipped by IMAP server: {exc}")

    for domain in PRIORITY_SENDER_DOMAINS:
        try:
            priority_uids.update(
                uid_search(
                    mail,
                    f"(FROM {imap_quote(domain)})",
                    retry_attempts=retry_attempts,
                    retry_sleep=retry_sleep,
                    label=f"IMAP sender search {domain}",
                )
            )
        except Exception as exc:
            print(f"Sample Calibration filter warning: sender domain '{domain}' skipped by IMAP server: {exc}")

    attachment_uids = (
        search_attachment_candidate_uids(
            mail,
            date_uids - subject_uids,
            retry_attempts=retry_attempts,
            retry_sleep=retry_sleep,
            max_probe=max_attachment_probe,
        )
        if enable_attachment_probe
        else set()
    )

    priority_date_uids = date_uids & priority_uids
    candidate_uids = (date_uids & subject_uids) | priority_date_uids | attachment_uids
    priority_sorted = sorted(candidate_uids & priority_uids, key=uid_sort_key)
    other_sorted = sorted(candidate_uids - priority_uids, key=uid_sort_key)
    ordered = sorted(candidate_uids, key=uid_sort_key, reverse=True)
    if limit:
        priority_latest = list(reversed(priority_sorted[-limit:]))
        remaining = max(limit - len(priority_latest), 0)
        other_latest = list(reversed(other_sorted[-remaining:])) if remaining else []
        ordered = priority_latest + other_latest

    print(
        "Sample Calibration IMAP filter: "
        f"date_matches={len(date_uids)}, subject_matches={len(subject_uids)}, "
        f"attachment_matches={len(attachment_uids)}, priority_sender_matches={len(priority_uids)}, "
        f"priority_date_matches={len(priority_date_uids)}, "
        f"candidates={len(candidate_uids)}"
    )
    return ordered


def write_invoice_output(path: Path, fields: dict[str, str], payload: bytes | None = None, page_number: int | None = None) -> tuple[Path, str]:
    required = ("invoice_date", "seller", "amount")
    if any(not fields.get(key) for key in required):
        target_dir = MANUAL_DIR
        status = "Need_Review"
    else:
        target_dir = PROCESSED_DIR / safe_name(fields.get("category", "其他"), "其他")
        status = "Parsed"
    target_dir.mkdir(parents=True, exist_ok=True)
    date = fields.get("invoice_date", "未知日期").replace("/", "-")
    seller = safe_name(fields.get("seller", "未知开票方"))
    amount = safe_name(fields.get("amount", "0.00"))
    number = safe_name(fields.get("invoice_number", "无号码"))
    page_suffix = f"_p{page_number:03d}" if page_number else ""
    target = target_dir / f"{date}_{seller}_{amount}_{number}{page_suffix}{path.suffix.lower()}"
    counter = 2
    while target.exists():
        target = target_dir / f"{date}_{seller}_{amount}_{number}{page_suffix}_{counter}{path.suffix.lower()}"
        counter += 1
    if payload is not None:
        target.write_bytes(payload)
    else:
        shutil.copy2(path, target)
    return target, status


def copy_to_manual_subfolder(path: Path, subfolder: str) -> Path:
    target_dir = MANUAL_DIR / subfolder
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / path.name
    counter = 2
    while target.exists():
        target = target_dir / f"{path.stem}_{counter}{path.suffix}"
        counter += 1
    shutil.copy2(path, target)
    return target


def blank_manifest_row(
    *,
    status: str,
    subject: str,
    sender: str,
    sent_at: str,
    received_at: str,
    uid_text: str,
    links: list[str] | str = "",
    note: str = "",
    acquisition_method: str = "",
    link_platform: str = "",
    link_status: str = "",
) -> dict[str, str]:
    return {
        "status": status,
        "parse_status": "",
        "parse_engine": "",
        "page_number": "",
        "page_count": "",
        "mail_subject": subject,
        "mail_from": sender,
        "mail_date": sent_at,
        "mail_received_at": received_at,
        "source_uid": uid_text,
        "attachment_original_name": "",
        "original_file": "",
        "output_file": "",
        "sha256": "",
        "invoice_date": "",
        "seller": "",
        "purchaser": "",
        "amount": "",
        "invoice_code": "",
        "invoice_number": "",
        "invoice_kind": "",
        "category": "",
        "order_number": "",
        "relation_type": "",
        "relation_group": "",
        "related_invoice_number": "",
        "include_in_summary": "",
        "effective_amount": "",
        "links": " | ".join(links) if isinstance(links, list) else links,
        "note": note,
        "acquisition_method": acquisition_method,
        "link_platform": link_platform,
        "link_status": link_status,
    }


def rows_for_saved_artifact(
    saved: Path,
    attachment_original_name: str,
    *,
    subject: str,
    sender: str,
    sent_at: str,
    received_at: str,
    uid_text: str,
    discovered_links: list[str],
    vision_extract: Callable[[list[str], str], dict[str, str]] | None,
    acquisition_method: str,
    link_platform: str = "",
    link_status: str = "",
    link_note: str = "",
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    if saved.suffix.lower() == ".pdf":
        pdf_bytes = saved.read_bytes()
        parsed_pages = parse_invoice_pdf_pages(
            pdf_bytes,
            filename=saved.name,
            vision_extract=vision_extract,
        )
        for parsed in parsed_pages:
            fields = {
                k: str(parsed.get(k, ""))
                for k in (
                    "invoice_date",
                    "seller",
                    "purchaser",
                    "amount",
                    "invoice_code",
                    "invoice_number",
                    "invoice_kind",
                    "category",
                    "order_number",
                    "relation_type",
                    "related_invoice_number",
                )
            }
            subject_hints = invoice_hints_from_subject(subject)
            for key, value in subject_hints.items():
                if value and not fields.get(key):
                    fields[key] = value
            if not fields.get("category"):
                if any(token in fields.get("seller", "") for token in ("餐饮", "饭店", "酒楼", "餐厅", "餐馆", "美食")):
                    fields["category"] = "餐饮"
                else:
                    fields["category"] = "其他"
            page_number = int(parsed.get("page_number") or 1)
            page_count = int(parsed.get("page_count") or len(parsed_pages) or 1)
            parse_status = str(parsed.get("status", ""))
            parse_engine = str(parsed.get("engine", ""))
            with OUTPUT_LOCK:
                if parse_status == "non_china_invoice":
                    output_path = copy_to_manual_subfolder(saved, "非中国发票")
                    status = "Non_CN_Invoice"
                elif parse_status == "non_invoice_document":
                    output_path = copy_to_manual_subfolder(saved, "非发票")
                    status = "Non_Invoice_Document"
                else:
                    page_payload = split_pdf_page_to_bytes(pdf_bytes, page_number - 1) if page_count > 1 else None
                    output_path, status = write_invoice_output(
                        saved,
                        fields,
                        payload=page_payload,
                        page_number=page_number if page_count > 1 else None,
                    )
                    if parse_engine == "vision_fallback" and parse_status == "parsed":
                        status = "AI_Verified"
            discovered_page_links = list(discovered_links)
            discovered_page_links.extend(str(link) for link in parsed.get("links", []) if link)
            note = str(parsed.get("error") or parsed.get("model_error") or link_note or "")
            rows.append(
                {
                    "status": status,
                    "parse_status": parse_status,
                    "parse_engine": parse_engine,
                    "page_number": str(page_number),
                    "page_count": str(page_count),
                    "mail_subject": subject,
                    "mail_from": sender,
                    "mail_date": sent_at,
                    "mail_received_at": received_at,
                    "source_uid": uid_text,
                    "attachment_original_name": attachment_original_name,
                    "original_file": str(saved),
                    "output_file": str(output_path),
                    "sha256": file_sha256(output_path),
                    "invoice_date": fields.get("invoice_date", ""),
                    "seller": fields.get("seller", ""),
                    "purchaser": fields.get("purchaser", ""),
                    "amount": fields.get("amount", ""),
                    "invoice_code": fields.get("invoice_code", ""),
                    "invoice_number": fields.get("invoice_number", ""),
                    "invoice_kind": fields.get("invoice_kind", ""),
                    "category": fields.get("category", ""),
                    "order_number": fields.get("order_number", ""),
                    "relation_type": fields.get("relation_type", ""),
                    "relation_group": "",
                    "related_invoice_number": fields.get("related_invoice_number", ""),
                    "include_in_summary": "",
                    "effective_amount": "",
                    "links": " | ".join(dict.fromkeys(discovered_page_links[:20])),
                    "note": note,
                    "acquisition_method": acquisition_method,
                    "link_platform": link_platform,
                    "link_status": link_status,
                }
        )
        return rows

    if saved.suffix.lower() in {".xml", ".ofd"}:
        parsed = parse_structured_invoice_artifact(saved)
        fields = {
            k: str(parsed.get(k, ""))
            for k in (
                "invoice_date",
                "seller",
                "purchaser",
                "amount",
                "invoice_code",
                "invoice_number",
                "invoice_kind",
                "category",
                "order_number",
                "relation_type",
                "related_invoice_number",
            )
        }
        subject_hints = invoice_hints_from_subject(subject)
        for key, value in subject_hints.items():
            if value and not fields.get(key):
                fields[key] = value
        if not fields.get("category"):
            fields["category"] = categorize_structured_invoice(fields.get("seller", ""), subject)
        parse_status = str(parsed.get("status", "need_review"))
        parse_engine = str(parsed.get("engine", f"{saved.suffix.lower().lstrip('.')}_structured"))
        with OUTPUT_LOCK:
            output_path, status = write_invoice_output(saved, fields)
        note = str(parsed.get("error") or link_note or "")
        rows.append(
            {
                "status": status,
                "parse_status": parse_status,
                "parse_engine": parse_engine,
                "page_number": "1",
                "page_count": "1",
                "mail_subject": subject,
                "mail_from": sender,
                "mail_date": sent_at,
                "mail_received_at": received_at,
                "source_uid": uid_text,
                "attachment_original_name": attachment_original_name,
                "original_file": str(saved),
                "output_file": str(output_path),
                "sha256": file_sha256(output_path),
                "invoice_date": fields.get("invoice_date", ""),
                "seller": fields.get("seller", ""),
                "purchaser": fields.get("purchaser", ""),
                "amount": fields.get("amount", ""),
                "invoice_code": fields.get("invoice_code", ""),
                "invoice_number": fields.get("invoice_number", ""),
                "invoice_kind": fields.get("invoice_kind", ""),
                "category": fields.get("category", ""),
                "order_number": fields.get("order_number", ""),
                "relation_type": fields.get("relation_type", ""),
                "relation_group": "",
                "related_invoice_number": fields.get("related_invoice_number", ""),
                "include_in_summary": "",
                "effective_amount": "",
                "links": " | ".join(dict.fromkeys(discovered_links[:20])),
                "note": note,
                "acquisition_method": acquisition_method,
                "link_platform": link_platform,
                "link_status": link_status,
            }
        )
        return rows

    rows.append(
        {
            "status": "Downloaded_Raw_Attachment" if acquisition_method == "邮件附件" else "Link_Downloaded_Raw",
            "parse_status": "",
            "parse_engine": "",
            "page_number": "",
            "page_count": "",
            "mail_subject": subject,
            "mail_from": sender,
            "mail_date": sent_at,
            "mail_received_at": received_at,
            "source_uid": uid_text,
            "attachment_original_name": attachment_original_name,
            "original_file": str(saved),
            "output_file": str(saved),
            "sha256": file_sha256(saved),
            "invoice_date": "",
            "seller": "",
            "purchaser": "",
            "amount": "",
            "invoice_code": "",
            "invoice_number": "",
            "invoice_kind": "",
            "category": "",
            "order_number": "",
            "relation_type": "",
            "relation_group": "",
            "related_invoice_number": "",
            "include_in_summary": "",
            "effective_amount": "",
            "links": " | ".join(dict.fromkeys(discovered_links[:20])),
            "note": link_note,
            "acquisition_method": acquisition_method,
            "link_platform": link_platform,
            "link_status": link_status,
        }
    )
    return rows


def get_thread_mail(
    env: dict[str, str],
    mailbox: str,
    context: ssl.SSLContext,
    retry_attempts: int,
    retry_sleep: float,
    connections: list[imaplib.IMAP4_SSL],
    connections_lock: threading.Lock,
) -> imaplib.IMAP4_SSL:
    mail = getattr(THREAD_LOCAL, "mail", None)
    if mail is not None:
        return mail

    imap_host = env.get("QQ_IMAP_HOST", "imap.qq.com")
    imap_port = int(env.get("QQ_IMAP_PORT", "993"))
    imap_timeout = float(env.get("IMAP_TIMEOUT_SECONDS", "60"))
    email_address = env.get("QQ_EMAIL", "")
    auth_code = env.get("QQ_IMAP_AUTH_CODE", "")
    mail = retry_call(
        lambda: imaplib.IMAP4_SSL(imap_host, imap_port, ssl_context=context, timeout=imap_timeout),
        attempts=retry_attempts,
        sleep_seconds=retry_sleep,
        label="IMAP worker connect",
    )
    set_imap_socket_timeout(mail, imap_timeout)
    retry_call(
        lambda: mail.login(email_address, auth_code),
        attempts=retry_attempts,
        sleep_seconds=retry_sleep,
        label="IMAP worker login",
    )
    retry_call(
        lambda: mail.select(mailbox, readonly=True),
        attempts=retry_attempts,
        sleep_seconds=retry_sleep,
        label="IMAP worker select",
    )
    THREAD_LOCAL.mail = mail
    with connections_lock:
        connections.append(mail)
    return mail


def set_imap_socket_timeout(mail: imaplib.IMAP4_SSL, timeout_seconds: float) -> None:
    sock = getattr(mail, "sock", None)
    if sock is None:
        return
    try:
        sock.settimeout(timeout_seconds)
    except Exception:
        pass


def close_worker_connections(connections: list[imaplib.IMAP4_SSL]) -> None:
    for mail in connections:
        try:
            mail.close()
        except Exception:
            pass
        try:
            mail.logout()
        except Exception:
            pass


def date_range_cache_path(since: str, until: str, limit: int | None = None) -> Path:
    suffix = f"{since.replace('-', '')}_{until.replace('-', '')}"
    if limit:
        suffix = f"{suffix}_limit{limit}"
    return STATE_DIR / f"rows_{suffix}.jsonl"


def append_row_cache(cache_path: Path, mailbox: str, uid: bytes | str, status: str, rows: list[dict[str, str]], error: str = "") -> None:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    uid_text = uid.decode("ascii", errors="replace") if isinstance(uid, bytes) else str(uid)
    record = {
        "mailbox": mailbox,
        "uid": uid_text,
        "status": status,
        "records_count": len(rows),
        "processed_at": dt.datetime.now().isoformat(timespec="seconds"),
        "error": error[:500],
        "rows": rows,
    }
    with cache_path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def load_row_cache(cache_path: Path) -> list[dict[str, str]]:
    if not cache_path.exists():
        return []
    latest_by_uid: dict[tuple[str, str], dict[str, object]] = {}
    with cache_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            key = (str(record.get("mailbox") or ""), str(record.get("uid") or ""))
            latest_by_uid[key] = record
    rows: list[dict[str, str]] = []
    for record in latest_by_uid.values():
        for row in record.get("rows") or []:
            if isinstance(row, dict):
                rows.append({str(k): "" if v is None else str(v) for k, v in row.items()})
    return rows


def process_message_uid(
    env: dict[str, str],
    mailbox: str,
    message_uid: bytes,
    run_stamp: str,
    retry_attempts: int,
    retry_sleep: float,
    context: ssl.SSLContext,
    vision_extract: Callable[[list[str], str], dict[str, str]] | None,
    qr_url_extract: Callable[[list[str]], str] | None,
    connections: list[imaplib.IMAP4_SSL],
    connections_lock: threading.Lock,
) -> tuple[bytes, str, list[dict[str, str]], str]:
    message_rows: list[dict[str, str]] = []
    uid_text = message_uid.decode("ascii", errors="replace")
    try:
        mail = get_thread_mail(env, mailbox, context, retry_attempts, retry_sleep, connections, connections_lock)
        typ, header_data = retry_call(
            lambda uid=message_uid: mail.uid("fetch", uid, "(BODY.PEEK[HEADER] INTERNALDATE)"),
            attempts=retry_attempts,
            sleep_seconds=retry_sleep,
            label=f"IMAP header fetch UID {uid_text}",
        )
        if typ != "OK" or not header_data or not header_data[0]:
            return message_uid, "fetch_failed", message_rows, "IMAP header fetch failed"

        header_bytes = b""
        for item in header_data:
            if isinstance(item, tuple) and isinstance(item[1], bytes):
                header_bytes = item[1]
                break
        header_msg = email.message_from_bytes(header_bytes)
        subject = decode_mime_words(header_msg.get("Subject"))
        sender = decode_mime_words(header_msg.get("From"))
        sent_at = parse_date(header_msg.get("Date"))
        received_at = parse_internal_date(header_data[0][0], sent_at)
        subject_related = subject_has_invoice_keyword(subject)
        if not (subject_related or looks_invoice_related(subject, sender) or sender_is_priority(sender)):
            return message_uid, "skipped_subject_filter", message_rows, ""

        typ, fetched = retry_call(
            lambda uid=message_uid: mail.uid("fetch", uid, "(RFC822 INTERNALDATE)"),
            attempts=retry_attempts,
            sleep_seconds=retry_sleep,
            label=f"IMAP fetch UID {uid_text}",
        )
        if typ != "OK" or not fetched or not fetched[0]:
            return message_uid, "fetch_failed", message_rows, "IMAP fetch failed"

        msg = email.message_from_bytes(fetched[0][1])
        subject = decode_mime_words(msg.get("Subject")) or subject
        sender = decode_mime_words(msg.get("From")) or sender
        sent_at = parse_date(msg.get("Date")) or sent_at
        received_at = parse_internal_date(fetched[0][0], sent_at)
        subject_related = subject_has_invoice_keyword(subject)
        plain, html_text = get_text_parts(msg)
        body_text = f"{plain}\n{re.sub(r'<[^>]+>', ' ', html_text or '')}"
        attachment_related = message_has_invoice_attachment(msg)
        related = subject_related or attachment_related or looks_invoice_related(subject, sender, plain, html_text)
        if not related:
            return message_uid, "skipped_subject_filter", message_rows, ""

        if is_bank_statement_email(subject, sender, body_text):
            return message_uid, "skipped_bank_statement", message_rows, ""
        if is_non_reimbursable_email(subject, sender, body_text):
            return message_uid, "skipped_non_reimbursable", message_rows, ""
        if is_deferred_platform_email(subject, sender, body_text):
            return message_uid, "skipped_deferred_platform", message_rows, ""

        link_candidates = select_invoice_link_candidates(
            extract_link_candidates(plain, html_text) if related else [],
            subject,
            sender,
            body_text,
        )
        urls = [candidate["url"] for candidate in link_candidates]
        prefix = f"{run_stamp}_uid{uid_text}"
        target_dir = RAW_DIR / run_stamp
        saved_any = False

        for part in msg.walk():
            if part.get_content_maintype() == "multipart":
                continue
            disposition = str(part.get("Content-Disposition") or "").lower()
            filename = part.get_filename()
            if "attachment" not in disposition and not filename:
                continue
            saved_info = save_attachment(part, target_dir, prefix)
            if not saved_info:
                continue
            saved, attachment_original_name = saved_info
            saved_any = True
            message_rows.extend(
                rows_for_saved_artifact(
                    saved,
                    attachment_original_name,
                    subject=subject,
                    sender=sender,
                    sent_at=sent_at,
                    received_at=received_at,
                    uid_text=uid_text,
                    discovered_links=list(urls),
                    vision_extract=vision_extract,
                    acquisition_method="邮件附件",
                )
            )

        link_artifacts: list[dict[str, str]] = []
        unresolved_links: list[dict[str, str]] = []
        if related and link_candidates:
            link_prefix = f"{prefix}_link"
            link_artifacts, unresolved_links = download_invoice_links(
                link_candidates,
                env,
                target_dir,
                link_prefix,
                retry_attempts,
                retry_sleep,
            )
            for artifact in link_artifacts:
                artifact_path = Path(artifact["path"])
                message_rows.extend(
                    rows_for_saved_artifact(
                        artifact_path,
                        artifact.get("original_name") or artifact_path.name,
                        subject=subject,
                        sender=sender,
                        sent_at=sent_at,
                        received_at=received_at,
                        uid_text=uid_text,
                        discovered_links=[artifact.get("source_url", ""), artifact.get("resolved_url", "")],
                        vision_extract=vision_extract,
                        acquisition_method="正文链接",
                        link_platform=artifact.get("platform", ""),
                        link_status=artifact.get("status", "Link_Downloaded"),
                    )
                )

        qr_artifacts: list[dict[str, str]] = []
        unresolved_qr: list[dict[str, str]] = []
        qr_candidates = collect_qr_image_candidates(msg, plain, html_text, env, retry_attempts, retry_sleep) if related else []
        if qr_candidates:
            qr_artifacts, unresolved_qr = recover_qr_invoice_links(
                qr_candidates,
                qr_url_extract,
                env,
                target_dir,
                f"{prefix}_qr",
                retry_attempts,
                retry_sleep,
                subject,
                sender,
            )
            for artifact in qr_artifacts:
                artifact_path = Path(artifact["path"])
                message_rows.extend(
                    rows_for_saved_artifact(
                        artifact_path,
                        artifact.get("original_name") or artifact_path.name,
                        subject=subject,
                        sender=sender,
                        sent_at=sent_at,
                        received_at=received_at,
                        uid_text=uid_text,
                        discovered_links=[artifact.get("source_url", ""), artifact.get("resolved_url", "")],
                        vision_extract=vision_extract,
                        acquisition_method="二维码识别",
                        link_platform=artifact.get("platform", ""),
                        link_status=artifact.get("status", "Link_Downloaded"),
                    )
                )
            if unresolved_qr and not qr_artifacts:
                for unresolved in unresolved_qr[:5]:
                    message_rows.append(
                        blank_manifest_row(
                            status="QR_Need_Manual",
                            subject=subject,
                            sender=sender,
                            sent_at=sent_at,
                            received_at=received_at,
                            uid_text=uid_text,
                            links=unresolved.get("url", ""),
                            note=unresolved.get("note", "二维码需要人工确认。"),
                            acquisition_method="二维码识别",
                            link_platform=unresolved.get("platform", ""),
                            link_status="QR_Need_Manual",
                        )
                    )

        if related and link_candidates and not link_artifacts and not qr_artifacts:
            if unresolved_links:
                for unresolved in unresolved_links[:5]:
                    message_rows.append(
                        blank_manifest_row(
                            status=unresolved.get("status", "Link_Need_Manual"),
                            subject=subject,
                            sender=sender,
                            sent_at=sent_at,
                            received_at=received_at,
                            uid_text=uid_text,
                            links=unresolved.get("url", ""),
                            note=unresolved.get("note", "链接需要人工确认。"),
                            acquisition_method="正文链接",
                            link_platform=unresolved.get("platform", ""),
                            link_status=unresolved.get("status", "Link_Need_Manual"),
                        )
                    )
            elif not saved_any and not unresolved_qr:
                message_rows.append(
                    blank_manifest_row(
                        status="Link_Need_Manual",
                        subject=subject,
                        sender=sender,
                        sent_at=sent_at,
                        received_at=received_at,
                        uid_text=uid_text,
                        links=urls[:20],
                        note="邮件疑似发票，但链接未能自动解析出 PDF/OFD/XML，需要人工确认。",
                        acquisition_method="正文链接",
                        link_platform=", ".join(sorted({candidate.get("platform", "") for candidate in link_candidates if candidate.get("platform")})),
                        link_status="Link_Need_Manual",
                    )
                )
        return message_uid, "processed", message_rows, ""
    except Exception as exc:
        return message_uid, "failed", message_rows, str(exc)


def scan_mailbox(
    env: dict[str, str],
    since: str,
    until: str,
    limit: int | None,
    log_db: Path = DEFAULT_LOG_DB,
    reprocess: bool = False,
    row_cache: Path | None = None,
) -> list[dict[str, str]]:
    email_address = env.get("QQ_EMAIL", "")
    auth_code = env.get("QQ_IMAP_AUTH_CODE", "")
    if not email_address or not auth_code:
        raise SystemExit("Missing QQ_EMAIL or QQ_IMAP_AUTH_CODE in env file.")

    imap_host = env.get("QQ_IMAP_HOST", "imap.qq.com")
    imap_port = int(env.get("QQ_IMAP_PORT", "993"))
    imap_timeout = float(env.get("IMAP_TIMEOUT_SECONDS", "60"))
    mailbox = env.get("QQ_MAILBOX", "INBOX")
    retry_attempts = int(env.get("RETRY_ATTEMPTS", "3"))
    retry_sleep = float(env.get("RETRY_SLEEP_SECONDS", "2"))
    context = ssl.create_default_context()
    mail = retry_call(
        lambda: imaplib.IMAP4_SSL(imap_host, imap_port, ssl_context=context, timeout=imap_timeout),
        attempts=retry_attempts,
        sleep_seconds=retry_sleep,
        label="IMAP connect",
    )
    set_imap_socket_timeout(mail, imap_timeout)
    retry_call(
        lambda: mail.login(email_address, auth_code),
        attempts=retry_attempts,
        sleep_seconds=retry_sleep,
        label="IMAP login",
    )
    log = ProcessLog(log_db)
    row_cache = row_cache or date_range_cache_path(since, until, limit)
    if reprocess and row_cache.exists():
        row_cache.unlink()
    try:
        retry_call(
            lambda: mail.select(mailbox, readonly=True),
            attempts=retry_attempts,
            sleep_seconds=retry_sleep,
            label="IMAP select",
        )
        message_ids = search_invoice_candidate_uids(
            mail,
            since=since,
            until=until,
            retry_attempts=retry_attempts,
            retry_sleep=retry_sleep,
            limit=limit,
            enable_attachment_probe=str(env.get("ENABLE_ATTACHMENT_PROBE", "1")).lower() not in {"0", "false", "no", "off"},
            max_attachment_probe=int(env.get("MAX_ATTACHMENT_PROBE_UIDS", str(MAX_ATTACHMENT_PROBE_UIDS))),
            search_mode=str(env.get("INVOICE_SEARCH_MODE", "filtered")),
        )

        rows: list[dict[str, str]] = []
        run_stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        vision_extract = build_mimo_vision_extractor(env)
        qr_url_extract = build_mimo_qr_url_extractor(env)
        pending_ids = [message_uid for message_uid in message_ids if reprocess or not log.is_processed(mailbox, message_uid)]
        skipped_count = len(message_ids) - len(pending_ids)
        print(f"Sample Calibration concurrency: max_workers={MAX_WORKERS}, pending={len(pending_ids)}, skipped={skipped_count}")

        worker_connections: list[imaplib.IMAP4_SSL] = []
        worker_connections_lock = threading.Lock()
        try:
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                futures = {
                    executor.submit(
                        process_message_uid,
                        env,
                        mailbox,
                        message_uid,
                        run_stamp,
                        retry_attempts,
                        retry_sleep,
                        context,
                        vision_extract,
                        qr_url_extract,
                        worker_connections,
                        worker_connections_lock,
                    ): message_uid
                    for message_uid in pending_ids
                }
                with make_progress(len(futures), "Sample Calibration") as progress:
                    for future in as_completed(futures):
                        message_uid = futures[future]
                        try:
                            uid, status, message_rows, error = future.result()
                        except Exception as exc:
                            uid = message_uid
                            status = "failed"
                            message_rows = []
                            error = str(exc)
                        rows.extend(message_rows)
                        log.mark(mailbox, uid, status, len(message_rows), error)
                        append_row_cache(row_cache, mailbox, uid, status, message_rows, error)
                        progress.update(1)
                        progress.set_postfix(records=len(rows), last=status)
        finally:
            close_worker_connections(worker_connections)
        cached_rows = load_row_cache(row_cache)
        return cached_rows if cached_rows else rows
    finally:
        log.close()
        try:
            mail.close()
        except Exception:
            pass
        try:
            mail.logout()
        except Exception:
            pass


def write_report(rows: list[dict[str, str]]) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    rows = clean_manifest_rows(rows)
    report_path = REPORT_DIR / f"发票台账_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    fieldnames = [
        "row_source",
        "account_id",
        "account_provider",
        "account_email",
        "status",
        "parse_status",
        "parse_engine",
        "page_number",
        "page_count",
        "mail_subject",
        "mail_from",
        "mail_date",
        "mail_received_at",
        "source_uid",
        "attachment_original_name",
        "original_file",
        "output_file",
        "sha256",
        "invoice_date",
        "seller",
        "purchaser",
        "amount",
        "invoice_code",
        "invoice_number",
        "invoice_kind",
        "category",
        "order_number",
        "relation_type",
        "relation_group",
        "related_invoice_number",
        "include_in_summary",
        "effective_amount",
        "acquisition_method",
        "link_platform",
        "link_status",
        "links",
        "note",
    ]
    with report_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return report_path


def write_xlsx_report(rows: list[dict[str, str]], csv_path: Path) -> Path:
    xlsx_path = csv_path.with_suffix(".xlsx")
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    internal_columns = [
        "row_source",
        "account_id",
        "account_provider",
        "account_email",
        "status",
        "parse_status",
        "parse_engine",
        "invoice_date",
        "seller",
        "purchaser",
        "amount",
        "invoice_code",
        "invoice_number",
        "invoice_kind",
        "category",
        "order_number",
        "relation_type",
        "relation_group",
        "related_invoice_number",
        "include_in_summary",
        "effective_amount",
        "page_number",
        "page_count",
        "mail_received_at",
        "mail_date",
        "attachment_original_name",
        "output_file",
        "original_file",
        "mail_from",
        "mail_subject",
        "acquisition_method",
        "link_platform",
        "link_status",
        "links",
        "note",
        "sha256",
        "source_uid",
    ]
    rows = clean_manifest_rows(rows)
    df = pd.DataFrame(rows)
    for col in internal_columns:
        if col not in df.columns:
            df[col] = ""
    df = df[internal_columns]
    df["amount_numeric"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0)
    df["effective_amount_numeric"] = pd.to_numeric(df["effective_amount"], errors="coerce").fillna(df["amount_numeric"])

    def output_basename(value: object, fallback: object) -> str:
        text = str(value or "")
        if text:
            return Path(text).name
        return str(fallback or "")

    df["ledger_file_name"] = [output_basename(row.output_file, row.attachment_original_name) for row in df.itertuples()]
    valid_df = df[df["status"].isin(["Parsed", "AI_Verified"])].copy()
    countable_df = valid_df[valid_df["include_in_summary"].fillna("是").replace("", "是") != "否"].copy()
    pending_df_source = df[~df["status"].isin(["Parsed", "AI_Verified"])].copy()

    def extract_subject_invoice_number(subject: object) -> str:
        match = re.search(r"发票号码[：:]?\s*([0-9A-Za-z-]+)", str(subject or ""))
        return match.group(1) if match else ""

    def extract_subject_amount(subject: object) -> str:
        match = re.search(r"发票金额[：:]?\s*([0-9,]+(?:\.[0-9]{1,2})?)", str(subject or ""))
        return f"{float(match.group(1).replace(',', '')):.2f}" if match else ""

    def extract_subject_order(subject: object) -> str:
        text = str(subject or "")
        patterns = (
            r"订单\s*([A-Za-z0-9-]{6,})",
            r"订单【([A-Za-z0-9-]{6,})】",
            r"invoice\s*\(#([A-Za-z0-9-]+)\)",
        )
        for pattern in patterns:
            match = re.search(pattern, text, flags=re.IGNORECASE)
            if match:
                return match.group(1).strip()
        return ""

    def pending_review_key(row: object) -> str:
        invoice_number = str(row.get("subject_invoice_number") or "")  # type: ignore[attr-defined]
        if invoice_number:
            return f"invoice:{invoice_number}"
        order_number = str(row.get("subject_order_number") or "")  # type: ignore[attr-defined]
        if order_number:
            return f"order:{order_number}"
        amount = str(row.get("subject_amount") or "")  # type: ignore[attr-defined]
        subject = str(row.get("mail_subject") or "")  # type: ignore[attr-defined]
        if amount:
            return f"subject_amount:{subject}|{amount}"
        return f"uid_subject:{row.get('source_uid') or ''}|{subject}"  # type: ignore[attr-defined]

    def useful_link(values: list[str]) -> str:
        noise_tokens = ("w3.org", "baidu.com", "bilibili.com", "github.com", "nuonuo.com/", "nnwwxicon", "nnwzfblogo", "smartCode")
        urls: list[str] = []
        for value in values:
            for url in re.findall(r"https?://[^\s|]+", str(value or "")):
                if url not in urls:
                    urls.append(url)
        for url in urls:
            if not any(token.lower() in url.lower() for token in noise_tokens):
                return url
        return ""

    pending_records: list[dict[str, object]] = []
    if not pending_df_source.empty:
        pending_df_source["subject_invoice_number"] = pending_df_source["mail_subject"].map(extract_subject_invoice_number)
        pending_df_source["subject_amount"] = pending_df_source["mail_subject"].map(extract_subject_amount)
        pending_df_source["subject_order_number"] = pending_df_source["mail_subject"].map(extract_subject_order)
        pending_df_source["review_key"] = pending_df_source.apply(pending_review_key, axis=1)
        for _, group in pending_df_source.groupby("review_key", dropna=False):
            first = group.iloc[0]
            statuses = " / ".join(sorted(set(str(value) for value in group["status"] if value)))
            link_values = [str(value) for value in group["links"] if str(value or "")]
            notes = [str(value) for value in group["note"] if str(value or "")]
            judgement = "已标注：下载结果非发票" if set(group["status"]) == {"Non_Invoice_Link"} else "待下载/待解析"
            pending_records.append(
                {
                    "状态": statuses,
                    "判断": judgement,
                    "邮件主题": first.get("mail_subject", ""),
                    "发票号码线索": first.get("subject_invoice_number", ""),
                    "金额线索": first.get("subject_amount", ""),
                    "邮箱账号": first.get("account_id", ""),
                    "邮件接收时间": first.get("mail_received_at", ""),
                    "发件人": first.get("mail_from", ""),
                    "获取方式": " / ".join(sorted(set(str(value) for value in group["acquisition_method"] if value))),
                    "链接平台": " / ".join(sorted(set(str(value) for value in group["link_platform"] if value))),
                    "首选链接": useful_link(link_values),
                    "线索行数": len(group),
                    "备注": "；".join(dict.fromkeys(notes[:3])),
                    "邮件UID": " / ".join(dict.fromkeys(str(value) for value in group["source_uid"] if value)),
                }
            )
    pending_review_df = pd.DataFrame(
        pending_records,
        columns=["状态", "判断", "邮件主题", "发票号码线索", "金额线索", "邮箱账号", "邮件接收时间", "发件人", "获取方式", "链接平台", "首选链接", "线索行数", "备注", "邮件UID"],
    )

    ledger_columns = [
        ("row_source", "数据来源"),
        ("account_id", "邮箱账号"),
        ("account_provider", "邮箱类型"),
        ("account_email", "邮箱地址"),
        ("category", "发票类型"),
        ("invoice_date", "开票日期"),
        ("purchaser", "付款方"),
        ("seller", "收款方"),
        ("amount", "金额"),
        ("effective_amount", "计入金额"),
        ("include_in_summary", "计入汇总"),
        ("relation_type", "业务处理"),
        ("relation_group", "关联组"),
        ("related_invoice_number", "关联发票号码"),
        ("order_number", "订单/业务号"),
        ("ledger_file_name", "发票名"),
        ("invoice_kind", "增值税类型"),
        ("status", "状态"),
        ("acquisition_method", "获取方式"),
        ("link_platform", "链接平台"),
        ("link_status", "链接处理状态"),
        ("mail_received_at", "邮件接收时间"),
        ("attachment_original_name", "附件原名"),
        ("invoice_number", "发票号码"),
        ("invoice_code", "发票代码"),
        ("page_number", "页码"),
        ("output_file", "整理后文件"),
        ("original_file", "原始文件"),
        ("mail_from", "发件人"),
        ("mail_subject", "邮件主题"),
        ("links", "链接线索"),
        ("note", "备注"),
        ("source_uid", "邮件UID"),
    ]
    ledger_df = pd.DataFrame({header: valid_df[source] for source, header in ledger_columns})
    detail_df = pd.DataFrame({header: df[source] for source, header in ledger_columns})

    category_summary = (
        countable_df.groupby("category", dropna=False)
        .agg(records=("status", "count"), total_amount=("effective_amount_numeric", "sum"), ai_verified=("status", lambda s: int((s == "AI_Verified").sum())))
        .reset_index()
        .rename(columns={"category": "发票类型", "records": "张数", "total_amount": "金额合计", "ai_verified": "AI复核张数"})
        .sort_values("金额合计", ascending=False)
        if not countable_df.empty
        else pd.DataFrame(columns=["发票类型", "张数", "金额合计", "AI复核张数"])
    )
    kind_summary = (
        countable_df.groupby("invoice_kind", dropna=False)
        .agg(records=("status", "count"), total_amount=("effective_amount_numeric", "sum"))
        .reset_index()
        .rename(columns={"invoice_kind": "增值税类型", "records": "张数", "total_amount": "金额合计"})
        .sort_values("金额合计", ascending=False)
        if not countable_df.empty
        else pd.DataFrame(columns=["增值税类型", "张数", "金额合计"])
    )
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        ledger_df.to_excel(writer, sheet_name="发票台账", index=False)
        pending_review_df.to_excel(writer, sheet_name="待确认线索", index=False)
        category_summary.to_excel(writer, sheet_name="分类汇总", index=False)
        kind_summary.to_excel(writer, sheet_name="增值税类型汇总", index=False)
        detail_df.to_excel(writer, sheet_name="抓取明细", index=False)
        workbook = writer.book
        for sheet_name in ("发票台账", "待确认线索", "分类汇总", "增值税类型汇总", "抓取明细"):
            ws = workbook[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in column_cells[:200])
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 10), 48)
        workbook["抓取明细"].sheet_state = "hidden"
        ledger = workbook["发票台账"]
        ledger.freeze_panes = "A2"
        amount_col = list(ledger_df.columns).index("金额") + 1
        effective_amount_col = list(ledger_df.columns).index("计入金额") + 1
        status_col = list(ledger_df.columns).index("状态") + 1
        include_col = list(ledger_df.columns).index("计入汇总") + 1
        relation_col = list(ledger_df.columns).index("业务处理") + 1
        yellow_fill = PatternFill("solid", fgColor="FFF2CC")
        red_fill = PatternFill("solid", fgColor="F4CCCC")
        gray_fill = PatternFill("solid", fgColor="D9EAD3")
        relation_fill = PatternFill("solid", fgColor="E2F0D9")
        reversal_fill = PatternFill("solid", fgColor="FCE4D6")
        for row in range(2, ledger.max_row + 1):
            amount_cell = ledger.cell(row=row, column=amount_col)
            try:
                amount_cell.value = float(str(amount_cell.value or "").replace(",", ""))
            except ValueError:
                pass
            amount_cell.number_format = '#,##0.00'
            effective_cell = ledger.cell(row=row, column=effective_amount_col)
            try:
                effective_cell.value = float(str(effective_cell.value or "").replace(",", ""))
            except ValueError:
                pass
            effective_cell.number_format = '#,##0.00'
            value = ledger.cell(row=row, column=status_col).value
            relation_value = str(ledger.cell(row=row, column=relation_col).value or "")
            if value == "AI_Verified":
                for col in range(1, ledger.max_column + 1):
                    ledger.cell(row=row, column=col).fill = yellow_fill
            if relation_value == "换开原票":
                for col in range(1, ledger.max_column + 1):
                    ledger.cell(row=row, column=col).fill = gray_fill
                ledger.cell(row=row, column=include_col).fill = relation_fill
            elif relation_value == "换开后有效发票":
                ledger.cell(row=row, column=relation_col).fill = relation_fill
            elif relation_value == "红字冲销":
                for col in range(1, ledger.max_column + 1):
                    ledger.cell(row=row, column=col).fill = reversal_fill
            elif value in {"Need_Review", "Link_Need_Manual", "QR_Need_Manual", "Non_Invoice_Link"}:
                ledger.cell(row=row, column=status_col).fill = red_fill
            elif value == "Non_CN_Invoice":
                ledger.cell(row=row, column=status_col).fill = gray_fill
        summary_sheet = workbook["分类汇总"]
        if summary_sheet.max_row >= 2:
            total_amount_col = 3
            for row in range(2, summary_sheet.max_row + 1):
                summary_sheet.cell(row=row, column=total_amount_col).number_format = '#,##0.00'
        kind_sheet = workbook["增值税类型汇总"]
        if kind_sheet.max_row >= 2:
            for row in range(2, kind_sheet.max_row + 1):
                kind_sheet.cell(row=row, column=3).number_format = '#,##0.00'
    return xlsx_path


def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Collect 2026 invoices from QQ mailbox through read-only IMAP.")
    parser.add_argument("--env", type=Path, default=CONFIG_DIR / "invoice_mail.env")
    parser.add_argument("--since", default=None, help="Override START_DATE from env file.")
    parser.add_argument("--until", default=None, help="Override END_DATE from env file; defaults to today.")
    parser.add_argument("--limit", type=int, default=None, help="Only process latest N matched messages for a safe trial run.")
    parser.add_argument("--log-db", type=Path, default=DEFAULT_LOG_DB)
    parser.add_argument("--row-cache", type=Path, default=None, help="JSONL row cache for resumable ledger output.")
    parser.add_argument("--reprocess", action="store_true", help="Ignore UID checkpoint and process messages again.")
    args = parser.parse_args(list(argv) if argv is not None else None)

    for directory in (CONFIG_DIR, RAW_DIR, PROCESSED_DIR, MANUAL_DIR, REPORT_DIR, STATE_DIR):
        directory.mkdir(parents=True, exist_ok=True)
    env = load_env(args.env)
    started = time.time()
    if str(env.get("MIMO_CALIBRATION_MODE") or "").strip().lower() in {"1", "true", "yes", "on"}:
        print("Sample Calibration mode enabled.")
    since = args.since or env.get("START_DATE") or "2026-01-01"
    until = args.until or env.get("END_DATE") or dt.date.today().isoformat()
    row_cache = args.row_cache or date_range_cache_path(since, until, args.limit)
    rows = scan_mailbox(env, since, until, args.limit, log_db=args.log_db, reprocess=args.reprocess, row_cache=row_cache)
    report_path = write_report(rows)
    xlsx_path = write_xlsx_report(rows, report_path)
    print(
        json.dumps(
            {
                "status": "completed",
                "records": len(rows),
                "csv_report": str(report_path),
                "xlsx_report": str(xlsx_path),
                "row_cache": str(row_cache),
                "elapsed_seconds": round(time.time() - started, 1),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
